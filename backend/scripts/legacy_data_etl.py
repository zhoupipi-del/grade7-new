# -*- coding: utf-8 -*-
"""
legacy_data_etl.py — 旧系统数据清洗 ETL 脚本模板

实现 BOSS 要求的三条策略：
1. 影子存储策略 — 旧数据先入 legacy_db 视图，不直接进核心库
2. 按需激活 — 学生触发业务请求时才从 legacy_db 加载到新系统
3. 数据血缘标记 — 每条导入数据标记 sync_status + lineage_ref

使用方法：
    python legacy_data_etl.py --action=import-students --source=old_students.xlsx
    python legacy_data_etl.py --action=import-classes --source=old_classes.xlsx
    python legacy_data_etl.py --action=import-teachers --source=old_teachers.xlsx
    python legacy_data_etl.py --action=activate-student --student-no=202670101
    python legacy_data_etl.py --action=verify --batch-id=1
"""

import argparse
import json
import logging
import os
import sys
from datetime import datetime, date
from typing import Optional, List, Dict, Any
from dataclasses import dataclass, field

# 日志配置
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("legacy_etl.log", encoding="utf-8"),
        logging.StreamHandler(),
    ],
)
logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════
# 数据血缘标记
# ═══════════════════════════════════════════════════════════════

@dataclass
class LineageMarker:
    """数据血缘标记 — 记录每条导入数据的来源信息"""
    source_system: str = "legacy_flask"
    source_table: str = ""
    source_id: str = ""
    target_table: str = ""
    target_id: str = ""
    batch_id: str = ""
    sync_status: str = "legacy"  # native / legacy / imported
    imported_at: str = field(default_factory=lambda: datetime.now().isoformat())
    field_mapping: Dict[str, str] = field(default_factory=dict)
    transform_notes: str = ""

    def to_dict(self) -> dict:
        return {k: v for k, v in self.__dict__.items()}

    def to_lineage_ref(self) -> str:
        """生成血缘引用ID"""
        return f"{self.source_system}:{self.source_table}:{self.source_id}"


# ═══════════════════════════════════════════════════════════════
# 影子存储 — 旧数据的临时存储层
# ═══════════════════════════════════════════════════════════════

class ShadowStore:
    """
    影子存储 — 旧数据先存到这里，不直接进 WINGS 核心库。
    当学生触发业务请求时，才从影子存储按需激活到新系统。

    实际实现可以用：
    - 独立的 SQLite/MySQL 数据库
    - JSON 文件
    - Redis 缓存
    这里用 JSON 文件作为模板示例。
    """

    def __init__(self, storage_dir: str = "legacy_shadow"):
        self.storage_dir = storage_dir
        os.makedirs(storage_dir, exist_ok=True)

    def save_students(self, students: List[dict], batch_id: str) -> str:
        """保存旧学生数据到影子存储"""
        path = os.path.join(self.storage_dir, f"students_{batch_id}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump({
                "batch_id": batch_id,
                "total": len(students),
                "imported_at": datetime.now().isoformat(),
                "students": students,
            }, f, ensure_ascii=False, indent=2)
        logger.info(f"影子存储: {len(students)} 条学生数据保存到 {path}")
        return path

    def save_classes(self, classes: List[dict], batch_id: str) -> str:
        path = os.path.join(self.storage_dir, f"classes_{batch_id}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump({
                "batch_id": batch_id,
                "total": len(classes),
                "imported_at": datetime.now().isoformat(),
                "classes": classes,
            }, f, ensure_ascii=False, indent=2)
        logger.info(f"影子存储: {len(classes)} 条班级数据保存到 {path}")
        return path

    def save_teachers(self, teachers: List[dict], batch_id: str) -> str:
        path = os.path.join(self.storage_dir, f"teachers_{batch_id}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump({
                "batch_id": batch_id,
                "total": len(teachers),
                "imported_at": datetime.now().isoformat(),
                "teachers": teachers,
            }, f, ensure_ascii=False, indent=2)
        logger.info(f"影子存储: {len(teachers)} 条教师数据保存到 {path}")
        return path

    def load_students(self, batch_id: str) -> List[dict]:
        path = os.path.join(self.storage_dir, f"students_{batch_id}.json")
        if not os.path.exists(path):
            return []
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("students", [])

    def find_student(self, keyword: str) -> Optional[dict]:
        """在影子存储中查找学生（按学号或姓名）"""
        for fname in os.listdir(self.storage_dir):
            if not fname.startswith("students_"):
                continue
            path = os.path.join(self.storage_dir, fname)
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            for s in data.get("students", []):
                if s.get("student_no") == keyword or s.get("name") == keyword:
                    return s
        return None


# ═══════════════════════════════════════════════════════════════
# ETL 管道 — 抽取/转换/加载
# ═══════════════════════════════════════════════════════════════

class ETLPipeline:
    """
    ETL 管道 — 从旧系统数据抽取、清洗转换、加载到影子存储。
    激活阶段才从影子存储加载到 WINGS 核心库。
    """

    def __init__(self, shadow_store: ShadowStore):
        self.shadow = shadow_store
        self.batch_id = datetime.now().strftime("%Y%m%d_%H%M%S")

    # ── 抽取 ──

    def extract_from_excel(self, file_path: str, sheet_name: str = "Sheet1") -> List[dict]:
        """从 Excel 抽取旧数据"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            headers = [cell.value for cell in ws[1]]

            data = []
            for row in rows:
                if not row[0]:
                    continue
                record = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        key = str(headers[i]).strip()
                        record[key] = val
                if record:
                    data.append(record)

            logger.info(f"抽取: {file_path} -> {len(data)} 条记录")
            return data
        except Exception as e:
            logger.error(f"抽取失败: {e}")
            raise

    def extract_from_csv(self, file_path: str, encoding: str = "utf-8") -> List[dict]:
        """从 CSV 抽取旧数据"""
        import csv
        data = []
        with open(file_path, "r", encoding=encoding) as f:
            reader = csv.DictReader(f)
            for row in reader:
                data.append(dict(row))
        logger.info(f"抽取: {file_path} -> {len(data)} 条记录")
        return data

    # ── 转换 ──

    def transform_students(self, raw_students: List[dict]) -> List[dict]:
        """
        清洗学生数据：
        - 字段名映射（旧系统 -> WINGS）
        - 性别枚举统一（男->M, 女->F）
        - 日期格式统一
        - 学号重新生成（保留旧学号映射）
        - 标记 sync_status=legacy
        """
        # 旧系统字段名 -> WINGS 字段名映射（需根据实际情况调整）
        field_mapping = {
            "姓名": "name",
            "学号": "student_no",
            "性别": "gender",
            "出生日期": "birth_date",
            "身份证号": "id_card",
            "民族": "nationality",
            "班级": "class_name",
            "班级ID": "class_id",
            "年级": "grade_name",
            "年级ID": "grade_id",
            "家庭住址": "address",
            "家长姓名": "parent1_name",
            "家长电话": "parent1_phone",
            "家长关系": "parent1_relation",
            "入学日期": "enrolled_at",
        }

        transformed = []
        errors = []

        for i, raw in enumerate(raw_students):
            try:
                # 字段映射
                record = {}
                for old_key, new_key in field_mapping.items():
                    if old_key in raw:
                        record[new_key] = raw[old_key]

                # 必填字段校验
                if not record.get("name"):
                    errors.append({"row": i + 2, "error": "姓名为空"})
                    continue

                # 性别转换
                gender = record.get("gender")
                if gender:
                    gender_str = str(gender).strip()
                    if gender_str in ("男", "M", "male"):
                        record["gender"] = "M"
                    elif gender_str in ("女", "F", "female"):
                        record["gender"] = "F"
                    else:
                        record["gender"] = None

                # 日期格式统一
                for date_field in ("birth_date", "enrolled_at"):
                    val = record.get(date_field)
                    if val and isinstance(val, str):
                        try:
                            record[date_field] = datetime.strptime(val[:10], "%Y-%m-%d").date().isoformat()
                        except ValueError:
                            try:
                                record[date_field] = datetime.strptime(val[:10], "%Y/%m/%d").date().isoformat()
                            except ValueError:
                                record[date_field] = None
                    elif val and isinstance(val, datetime):
                        record[date_field] = val.date().isoformat()
                    elif val and isinstance(val, date):
                        record[date_field] = val.isoformat()

                # 保留旧学号用于映射
                old_student_no = record.get("student_no")
                record["legacy_student_no"] = old_student_no

                # 血缘标记
                record["sync_status"] = "legacy"
                record["lineage_ref"] = f"legacy_flask:students:{old_student_no or i}"
                record["batch_id"] = self.batch_id

                transformed.append(record)
            except Exception as e:
                errors.append({"row": i + 2, "name": raw.get("姓名", ""), "error": str(e)})

        logger.info(f"转换: {len(raw_students)} -> {len(transformed)} 条 (失败 {len(errors)})")
        if errors:
            for err in errors[:10]:
                logger.warning(f"  转换失败: {err}")

        return transformed

    def transform_classes(self, raw_classes: List[dict]) -> List[dict]:
        """清洗班级数据"""
        field_mapping = {
            "班级名称": "name",
            "班级": "name",
            "年级": "grade_name",
            "年级ID": "grade_id",
            "班主任": "head_teacher_name",
            "班主任ID": "head_teacher_id",
            "人数": "student_count",
        }

        transformed = []
        for raw in raw_classes:
            record = {}
            for old_key, new_key in field_mapping.items():
                if old_key in raw:
                    record[new_key] = raw[old_key]
            if not record.get("name"):
                continue
            record["sync_status"] = "legacy"
            record["lineage_ref"] = f"legacy_flask:classes:{record.get('name')}"
            record["batch_id"] = self.batch_id
            transformed.append(record)

        logger.info(f"班级转换: {len(raw_classes)} -> {len(transformed)} 条")
        return transformed

    def transform_teachers(self, raw_teachers: List[dict]) -> List[dict]:
        """清洗教师数据"""
        field_mapping = {
            "姓名": "name",
            "工号": "employee_no",
            "性别": "gender",
            "任教学科": "subject",
            "科目": "subject",
            "职称": "title",
            "电话": "phone",
            "手机": "phone",
        }

        transformed = []
        for raw in raw_teachers:
            record = {}
            for old_key, new_key in field_mapping.items():
                if old_key in raw:
                    record[new_key] = raw[old_key]
            if not record.get("name"):
                continue
            # 性别转换
            gender = record.get("gender")
            if gender:
                g = str(gender).strip()
                record["gender"] = "M" if g in ("男", "M") else "F" if g in ("女", "F") else None
            record["sync_status"] = "legacy"
            record["lineage_ref"] = f"legacy_flask:teachers:{record.get('employee_no') or record.get('name')}"
            record["batch_id"] = self.batch_id
            transformed.append(record)

        logger.info(f"教师转换: {len(raw_teachers)} -> {len(transformed)} 条")
        return transformed

    # ── 加载到影子存储 ──

    def load_to_shadow(self, data: List[dict], data_type: str) -> str:
        """加载清洗后的数据到影子存储"""
        if data_type == "students":
            return self.shadow.save_students(data, self.batch_id)
        elif data_type == "classes":
            return self.shadow.save_classes(data, self.batch_id)
        elif data_type == "teachers":
            return self.shadow.save_teachers(data, self.batch_id)
        else:
            raise ValueError(f"未知数据类型: {data_type}")

    # ── 完整 ETL 流程 ──

    def run_full_etl(self, source_file: str, data_type: str) -> dict:
        """执行完整的 ETL 流程"""
        logger.info(f"=== ETL 开始: {data_type} from {source_file} ===")

        # 1. 抽取
        if source_file.endswith(".xlsx") or source_file.endswith(".xls"):
            raw_data = self.extract_from_excel(source_file)
        elif source_file.endswith(".csv"):
            raw_data = self.extract_from_csv(source_file)
        else:
            raise ValueError(f"不支持的文件格式: {source_file}")

        # 2. 转换
        if data_type == "students":
            transformed = self.transform_students(raw_data)
        elif data_type == "classes":
            transformed = self.transform_classes(raw_data)
        elif data_type == "teachers":
            transformed = self.transform_teachers(raw_data)
        else:
            raise ValueError(f"未知数据类型: {data_type}")

        # 3. 加载到影子存储
        shadow_path = self.load_to_shadow(transformed, data_type)

        # 4. 生成血缘报告
        lineage_report = {
            "batch_id": self.batch_id,
            "data_type": data_type,
            "source_file": source_file,
            "raw_count": len(raw_data),
            "transformed_count": len(transformed),
            "shadow_path": shadow_path,
            "imported_at": datetime.now().isoformat(),
            "markers": [
                {
                    "source_system": "legacy_flask",
                    "source_table": data_type,
                    "sync_status": "legacy",
                    "lineage_ref": item.get("lineage_ref", ""),
                }
                for item in transformed
            ],
        }

        report_path = os.path.join(
            self.shadow.storage_dir,
            f"lineage_report_{data_type}_{self.batch_id}.json",
        )
        with open(report_path, "w", encoding="utf-8") as f:
            json.dump(lineage_report, f, ensure_ascii=False, indent=2)

        logger.info(f"=== ETL 完成: {len(transformed)} 条 -> {shadow_path} ===")
        logger.info(f"血缘报告: {report_path}")

        return lineage_report

    # ── 按需激活 ──

    def activate_student(self, student_no: str) -> Optional[dict]:
        """
        按需激活 — 当学生在 WINGS 中触发业务请求时，
        从影子存储加载该学生的历史数据到新系统。
        """
        # 从影子存储查找
        legacy_data = self.shadow.find_student(student_no)
        if not legacy_data:
            logger.warning(f"影子存储中未找到学生: {student_no}")
            return None

        logger.info(f"按需激活学生: {student_no} -> WINGS")

        # 调用 WINGS API 创建学籍（标记 sync_status=legacy）
        # 实际实现需要通过 HTTP 调用 WINGS API
        # 这里返回清洗后的数据，由调用方通过 API 导入

        activation_record = {
            **legacy_data,
            "sync_status": "legacy",
            "activated_at": datetime.now().isoformat(),
            "lineage_ref": f"legacy_flask:students:{legacy_data.get('legacy_student_no', '')}",
        }

        return activation_record

    # ── 验证 ──

    def verify_import(self, batch_id: str) -> dict:
        """验证导入数据的完整性和正确性"""
        report_path = os.path.join(
            self.shadow.storage_dir,
            f"lineage_report_students_{batch_id}.json",
        )

        if not os.path.exists(report_path):
            return {"error": f"未找到批次报告: {batch_id}"}

        with open(report_path, "r", encoding="utf-8") as f:
            report = json.load(f)

        # 对比源数据和影子存储数据
        shadow_students = self.shadow.load_students(batch_id)

        verification = {
            "batch_id": batch_id,
            "source_count": report["raw_count"],
            "shadow_count": len(shadow_students),
            "match": report["raw_count"] == len(shadow_students),
            "markers_count": len(report.get("markers", [])),
            "all_marked_legacy": all(
                m.get("sync_status") == "legacy" for m in report.get("markers", [])
            ),
            "verified_at": datetime.now().isoformat(),
        }

        logger.info(f"验证结果: {verification}")
        return verification


# ═══════════════════════════════════════════════════════════════
# CLI 入口
# ═══════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="WINGS 旧数据 ETL 工具")
    parser.add_argument("--action", required=True,
                        choices=["import-students", "import-classes", "import-teachers",
                                 "activate-student", "verify"],
                        help="执行的操作")
    parser.add_argument("--source", help="源数据文件路径")
    parser.add_argument("--student-no", help="要激活的学生学号")
    parser.add_argument("--batch-id", help="要验证的批次ID")

    args = parser.parse_args()

    shadow = ShadowStore()
    etl = ETLPipeline(shadow)

    if args.action == "import-students":
        if not args.source:
            print("错误: 需要 --source 参数")
            sys.exit(1)
        report = etl.run_full_etl(args.source, "students")
        print(f"\n导入完成:")
        print(f"  源数据: {report['raw_count']} 条")
        print(f"  清洗后: {report['transformed_count']} 条")
        print(f"  影子存储: {report['shadow_path']}")
        print(f"  批次ID: {report['batch_id']}")

    elif args.action == "import-classes":
        if not args.source:
            print("错误: 需要 --source 参数")
            sys.exit(1)
        report = etl.run_full_etl(args.source, "classes")
        print(f"\n导入完成: {report['transformed_count']} 个班级")

    elif args.action == "import-teachers":
        if not args.source:
            print("错误: 需要 --source 参数")
            sys.exit(1)
        report = etl.run_full_etl(args.source, "teachers")
        print(f"\n导入完成: {report['transformed_count']} 个教师")

    elif args.action == "activate-student":
        if not args.student_no:
            print("错误: 需要 --student-no 参数")
            sys.exit(1)
        result = etl.activate_student(args.student_no)
        if result:
            print(f"\n学生激活成功:")
            print(f"  姓名: {result.get('name')}")
            print(f"  学号: {result.get('student_no')}")
            print(f"  旧学号: {result.get('legacy_student_no')}")
            print(f"  血缘: {result.get('lineage_ref')}")
        else:
            print(f"\n未找到学生: {args.student_no}")

    elif args.action == "verify":
        if not args.batch_id:
            print("错误: 需要 --batch-id 参数")
            sys.exit(1)
        result = etl.verify_import(args.batch_id)
        print(f"\n验证结果:")
        for k, v in result.items():
            print(f"  {k}: {v}")


if __name__ == "__main__":
    main()

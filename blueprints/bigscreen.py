"""全校德育数据大屏"""
from flask import Blueprint, render_template, request, jsonify, session
from datetime import datetime, timedelta
from collections import defaultdict
import json
from sqlalchemy import func, text

from decorators import login_required, require_role
from utils import get_local_now
from utils.statistics_service import (
    get_basic_stats, get_discipline_stats, get_attendance_stats,
    get_wing_stats, get_mental_health_stats, get_notice_read_rate,
    get_visit_stats, get_risk_stats, get_trend_data,
    get_class_score_ranking, get_score_overview,
)
from models import db, User, Student, Class, Grade, DisciplineRecord, RoutineScore, Task, WingsScore, HomeVisit, Notice, NoticeReceipt, Score, Exam, MentalHealthAssessment, RiskRecord, Attendance, Subject

bigscreen_bp = Blueprint("bigscreen", __name__, url_prefix="/bigscreen")


@bigscreen_bp.route("/")
@login_required
@require_role("ms_admin", "grade_leader")
def index():
    """大屏首页"""
    return render_template("bigscreen/index.html")


@bigscreen_bp.route("/data")
@login_required
@require_role("ms_admin", "grade_leader")
def data():
    """返回大屏需要的JSON数据"""
    # 从 session 中获取当前用户上下文 (项目使用 Flask session，非 Flask-Login)
    user_role = session.get("role")
    user_grade_id = None
    if user_role == 'grade_leader':
        user = User.query.get(session.get("user_id"))
        if user:
            user_grade_id = user.grade_id

    # ── 共享统计查询层 (全校视图, grade_id=None) ──
    basic = get_basic_stats()

    # 违纪统计 (近30天)
    since = get_local_now() - timedelta(days=30)
    disc = get_discipline_stats(since=since)

    # 五翼均值
    wing = get_wing_stats()

    # 班级得分排行
    class_scores_data = get_class_score_ranking()

    # 通知阅读率
    notice = get_notice_read_rate()

    # 家访统计
    visit = get_visit_stats()

    # 成绩概览
    score_overview = get_score_overview()

    # 心理健康风险分布
    mh = get_mental_health_stats()

    # 考勤概况
    att = get_attendance_stats(since=since)

    # 7天趋势
    trend = get_trend_data()

    # ==================== Phase 4 - 刀 1: 数据供给侧清洗与战略扩容 ====================
    # 1. 动态获取最新扫描日期，确保度量衡对齐最新批次
    latest_scan = RiskRecord.query.order_by(RiskRecord.scan_date.desc()).first()
    target_scan_date = latest_scan.scan_date if latest_scan else get_local_now().date()

    # 2. 角色安全守卫与年级边界隔离 (RiskRecord 自带 grade_id，无需穿透 Student)
    risk_query = RiskRecord.query.filter(RiskRecord.scan_date == target_scan_date)
    if user_role == 'grade_leader' and user_grade_id:
        risk_query = risk_query.filter(RiskRecord.grade_id == user_grade_id)

    current_risks = risk_query.all()

    # 3. 预加载 Student/Grade/Class 名称映射 (批量查询，斩断 N+1)
    risk_student_ids = list(set(r.student_id for r in current_risks))
    students = Student.query.filter(Student.id.in_(risk_student_ids)).all() if risk_student_ids else []
    student_map = {s.id: s for s in students}
    grade_map = {g.id: g.name for g in Grade.query.all()}
    class_map = {c.id: c.name for c in Class.query.all()}

    # 4. 内存多维聚合提取器 (单趟扫描，斩断 I/O 穿透)
    risk_counts = {"red": 0, "yellow": 0, "green": 0}
    type_distribution = defaultdict(int)
    high_risk_pool = []

    for record in current_risks:
        risk_counts[record.risk_level] = risk_counts.get(record.risk_level, 0) + 1

        # 解析 warning_details JSON 数组 → 提取 type 热力分布
        warning_types = []
        try:
            details = json.loads(record.warning_details) if record.warning_details else []
        except Exception:
            details = []

        for item in details:
            w_type = item.get("type", "")
            if w_type:
                type_distribution[w_type] += 1
                warning_types.append(w_type)

        w_count = record.warning_count or len(warning_types)

        # 收集红色/黄色高危学生，供 Top-10 穿透名单使用
        if record.risk_level in ["red", "yellow"]:
            stu_obj = student_map.get(record.student_id)
            if stu_obj:
                high_risk_pool.append({
                    "student_id": record.student_id,
                    "name": stu_obj.name,
                    "class_name": f"{grade_map.get(stu_obj.grade_id, '')}{class_map.get(stu_obj.class_id, '')}",
                    "level": record.risk_level,
                    "warning_count": w_count,
                    "details": ", ".join(warning_types) if warning_types else "常规触发"
                })

    # 5. Top-10 核心高危名单 (按风险等级+触发预警项数量降序)
    level_weight = {"red": 2, "yellow": 1, "green": 0}
    high_risk_pool.sort(key=lambda x: (level_weight.get(x["level"], 0), x["warning_count"]), reverse=True)
    risk_top_10 = high_risk_pool[:10]

    # 6. 补齐近 7 天预警趋势线
    seven_days_ago = target_scan_date - timedelta(days=6)
    hist_query = RiskRecord.query.filter(
        RiskRecord.scan_date >= seven_days_ago,
        RiskRecord.scan_date <= target_scan_date
    )
    if user_role == 'grade_leader' and user_grade_id:
        hist_query = hist_query.filter(RiskRecord.grade_id == user_grade_id)

    historical_risks = hist_query.all()
    trend_map = defaultdict(lambda: {"red": 0, "yellow": 0, "green": 0})
    for h_rec in historical_risks:
        date_str = h_rec.scan_date.strftime("%m-%d")
        if h_rec.risk_level in trend_map[date_str]:
            trend_map[date_str][h_rec.risk_level] += 1

    sorted_dates = sorted(list(trend_map.keys()))
    risk_trend = {
        "dates": sorted_dates,
        "red": [trend_map[d]["red"] for d in sorted_dates],
        "yellow": [trend_map[d]["yellow"] for d in sorted_dates],
        "green": [trend_map[d]["green"] for d in sorted_dates]
    }

    # 7. 同步更新原有 risk_red/yellow/green 变量 (对外接口完全兼容)
    risk_red = risk_counts["red"]
    risk_yellow = risk_counts["yellow"]
    risk_green = risk_counts["green"]

    # 考勤概况 — 已由 get_attendance_stats 计算

    # 7天趋势 — 已由 get_trend_data 计算

    return jsonify({
        "total_students": basic["total_students"],
        "total_classes": basic["total_classes"],
        "total_teachers": basic["total_teachers"],
        "discipline_count": disc["total_count"],
        "level_warning": disc["by_type"].get("warning", 0),
        "level_minor": disc["by_type"].get("minor", 0),
        "level_major": disc["by_type"].get("major", 0),
        "wing_avg": round(float(wing["avg"]), 1),
        "class_scores": class_scores_data,
        "notice_count": notice["notice_count"],
        "notice_read_rate": notice["read_rate"],
        "visit_count": visit["total_count"],
        "trend": trend,
        # 新增
        "score_avg": score_overview["avg_score"],
        "score_pass_rate": score_overview["pass_rate"],
        "latest_exam_name": score_overview["exam_name"],
        "mh_high": mh["high"],
        "mh_medium": mh["medium"],
        "mh_low": mh["low"],
        "risk_red": risk_red,
        "risk_yellow": risk_yellow,
        "risk_green": risk_green,
        "att_rate": att["attendance_rate"],
        "att_total": att["total"],
        # 🚀 Phase 4 增量核心武器库
        "risk_by_type": dict(type_distribution),
        "risk_top_n": risk_top_10,
        "risk_trend": risk_trend,
        "latest_scan_date": target_scan_date.strftime("%Y-%m-%d"),
    })


@bigscreen_bp.route("/academic-radar")
@login_required
@require_role("ms_admin", "grade_leader")
def academic_radar():
    """学情驾驶舱 API: 学业断崖雷达 + 班主任红黑榜"""
    from collections import defaultdict

    # ===== PART 1: 班主任红黑榜（评分覆盖率排行）=====
    all_classes = Class.query.filter_by(is_active=True).order_by(Class.name).all()
    class_ids = [c.id for c in all_classes]

    # 批量查每班 wings_scores 覆盖人数（排除admin，只算班主任+科任老师）
    coverage_rows = db.session.query(
        WingsScore.class_id,
        func.count(func.distinct(WingsScore.student_id))
    ).filter(
        WingsScore.class_id.in_(class_ids),
        WingsScore.scorer_id != 1  # 排除admin
    ).group_by(WingsScore.class_id).all()
    coverage_map = {cid: cnt for cid, cnt in coverage_rows}

    # 批量查每班学生总数
    student_count_rows = db.session.query(
        Student.class_id,
        func.count(Student.id)
    ).filter(Student.class_id.in_(class_ids), Student.is_active == True)\
     .group_by(Student.class_id).all()
    student_count_map = {cid: cnt for cid, cnt in student_count_rows}

    teacher_ranking = []
    for c in all_classes:
        total = student_count_map.get(c.id, 0)
        covered = coverage_map.get(c.id, 0)
        rate = round(covered / total * 100, 1) if total > 0 else 0
        teacher_ranking.append({
            "class_id": c.id,
            "class_name": c.name,
            "total_students": total,
            "covered_students": covered,
            "coverage_rate": rate,
        })
    teacher_ranking.sort(key=lambda x: x["coverage_rate"])

    # ===== PART 2: 学业断崖/复苏雷达 =====
    exams = Exam.query.filter_by(grade_id=1).order_by(Exam.exam_date.asc()).all()
    cliffs = []
    recoveries = []

    if len(exams) >= 2:
        exam_ids = [e.id for e in exams]
        exam_map = {e.id: e for e in exams}

        # 批量载入所有成绩 (9286条，一次查询)
        all_scores = Score.query.filter(Score.exam_id.in_(exam_ids)).all()

        # 预加载学生/班级/科目映射
        sid_set = {s.student_id for s in all_scores}
        students = Student.query.filter(Student.id.in_(list(sid_set))).all()
        student_map = {s.id: s for s in students}
        class_map = {c.id: c.name for c in all_classes}
        subject_map = {s.id: s.name for s in Subject.query.all()}

        # 聚合: {(student_id, subject_id): [(exam_id, score), ...]}
        timeline = defaultdict(list)
        for s in all_scores:
            timeline[(s.student_id, s.subject_id)].append((s.exam_id, s.score))

        for (sid, subj_id), records in timeline.items():
            if len(records) < 2:
                continue
            records.sort(key=lambda x: x[0])
            prev_eid, prev_score = records[-2]
            curr_eid, curr_score = records[-1]
            # 跳过缺考（0分可能是缺考，排除人为制造噪音）
            if prev_score == 0 or curr_score == 0:
                continue
            delta = round(curr_score - prev_score, 1)

            stu = student_map.get(sid)
            if not stu:
                continue
            entry = {
                "student_id": sid,
                "student_name": stu.name,
                "class_name": class_map.get(stu.class_id, "?"),
                "subject": subject_map.get(subj_id, f"科目{subj_id}"),
                "prev_exam": exam_map[prev_eid].name.replace("2025年", "").replace("2026年", "").replace("七年级", ""),
                "curr_exam": exam_map[curr_eid].name.replace("2025年", "").replace("2026年", "").replace("七年级", ""),
                "prev_score": round(prev_score, 1),
                "curr_score": round(curr_score, 1),
                "delta": delta,
                "delta_abs": abs(delta),
            }
            if delta <= -15:
                cliffs.append(entry)
            elif delta >= 20:
                recoveries.append(entry)

        cliffs.sort(key=lambda x: x["delta"])
        recoveries.sort(key=lambda x: x["delta"], reverse=True)
        cliffs = cliffs[:20]
        recoveries = recoveries[:20]

    return jsonify({
        "teacher_ranking": teacher_ranking,
        "cliffs": cliffs,
        "recoveries": recoveries,
        "total_exams": len(exams),
    })


# ══════════════════════════════════════════════════════════════
#  🚀 时空共时性死党图谱 — 图数据挖掘
# ══════════════════════════════════════════════════════════════
@bigscreen_bp.route("/social-graph")
@login_required
@require_role("ms_admin", "grade_leader")
def social_graph():
    """隐形违纪社交网络拓扑图 API
    算法：时空共时性 (Spatio-temporal Sync) — 
    同一30分钟窗口内被同一老师同时扣分的学生之间产生引力边
    """
    from collections import defaultdict
    import itertools
    from datetime import timedelta

    user_role = session.get("role")
    user_grade_id = None
    if user_role == 'grade_leader':
        user = User.query.get(session.get("user_id"))
        if user:
            user_grade_id = user.grade_id

    # ── 1. 主力信号源：WingsScore 教师评分共时性 ──
    wings_query = db.session.query(
        WingsScore.student_id,
        WingsScore.scorer_id,
        WingsScore.scorer_type,
        WingsScore.created_at,
        WingsScore.dimension,
        WingsScore.score
    )
    if user_grade_id:
        wings_query = wings_query.filter(WingsScore.grade_id == user_grade_id)

    wings_records = wings_query.all()

    # ── 2. 辅助信号源：DisciplineRecord（同一违纪事件多条记录）──
    disc_query = db.session.query(
        DisciplineRecord.student_id,
        DisciplineRecord.created_by,
        DisciplineRecord.created_at,
        DisciplineRecord.type
    )
    if user_grade_id:
        disc_query = disc_query.join(Student).filter(Student.grade_id == user_grade_id)

    disc_records = disc_query.all()

    # ── 3. 构建时间窗口哈希桶 ──
    time_windows = defaultdict(list)

    for r in wings_records:
        if r.created_at is None:
            continue
        # 30分钟粒度：year-month-day-hour-halfhour
        half = "00" if r.created_at.minute < 30 else "30"
        window_key = f"{r.scorer_id}|{r.created_at.strftime('%Y-%m-%d %H')}:{half}"
        time_windows[window_key].append({
            "student_id": r.student_id,
            "dimension": r.dimension,
            "score": r.score,
            "source": "wings"
        })

    for r in disc_records:
        if r.created_at is None:
            continue
        half = "00" if r.created_at.minute < 30 else "30"
        window_key = f"teacher_{r.created_by}|{r.created_at.strftime('%Y-%m-%d %H')}:{half}"
        time_windows[window_key].append({
            "student_id": r.student_id,
            "type": r.type,
            "source": "discipline"
        })

    # ── 4. 构建共现引力矩阵 ──
    edge_weights = defaultdict(int)
    node_weights = defaultdict(int)
    edge_details = defaultdict(list)  # 存储每次共现的详情

    for window_key, entries in time_windows.items():
        student_ids = list(set(e["student_id"] for e in entries))
        for sid in student_ids:
            node_weights[sid] += 1

        if len(student_ids) >= 2:
            for s1, s2 in itertools.combinations(sorted(student_ids), 2):
                edge_key = (min(s1, s2), max(s1, s2))
                edge_weights[edge_key] += 1
                # 保存共现详情（最多存5条）
                if len(edge_details[edge_key]) < 5:
                    scorer_part = window_key.split("|")[0]
                    time_part = window_key.split("|")[1]
                    edge_details[edge_key].append({
                        "time": time_part,
                        "scorer_context": scorer_part,
                        "entries": [
                            {"sid": e["student_id"],
                             "info": e.get("dimension", e.get("type", "?"))}
                            for e in entries if e["student_id"] in (s1, s2)
                        ][:4]
                    })

    # ── 5. 批量预加载学生/班级映射 ──
    all_sids = set(node_weights.keys())
    students = Student.query.filter(Student.id.in_(list(all_sids))).all() if all_sids else []
    student_map = {s.id: s for s in students}
    class_map = {c.id: c.name for c in Class.query.filter_by(is_active=True).all()}

    # ── 6. 组装 ECharts Graph 节点 ──
    nodes = []
    for sid, weight in sorted(node_weights.items(), key=lambda x: -x[1]):
        stu = student_map.get(sid)
        if not stu:
            continue
        # 节点大小映射：共振密度
        symbol_size = min(max(weight * 5, 18), 70)
        # 分类：共振≥5次=核心高危节点
        category = 1 if weight >= 5 else 0
        nodes.append({
            "id": str(sid),
            "name": stu.name,
            "symbolSize": symbol_size,
            "value": weight,
            "category": category,
            "itemStyle": {
                "color": "#ff4757" if category == 1 else "#ffa502"
            },
            "tooltip_extra": f"{class_map.get(stu.class_id, '?')} · 共振{weight}次"
        })

    # ── 7. 组装边（至少共现2次才视为死党）──
    links = []
    gang_warnings = []  # 高危团伙报警
    for (s1, s2), weight in sorted(edge_weights.items(), key=lambda x: -x[1]):
        if weight < 2:
            continue
        s1_stu = student_map.get(s1)
        s2_stu = student_map.get(s2)
        if not s1_stu or not s2_stu:
            continue

        line_width = min(weight * 1.5, 8)
        # 颜色：≥5次共振 = 红色警告边
        edge_color = "rgba(255,71,87,0.7)" if weight >= 5 else "rgba(255,165,2,0.4)"

        links.append({
            "source": str(s1),
            "target": str(s2),
            "value": weight,
            "lineStyle": {
                "width": line_width,
                "color": edge_color,
                "curveness": 0.1
            },
            "label": {
                "show": weight >= 5,
                "formatter": f"{weight}次"
            }
        })

        # 高危团伙判定：跨班+高频共振
        if weight >= 3 and s1_stu.class_id != s2_stu.class_id:
            gang_warnings.append({
                "student1": s1_stu.name,
                "student1_id": s1,
                "class1": class_map.get(s1_stu.class_id, "?"),
                "student2": s2_stu.name,
                "student2_id": s2,
                "class2": class_map.get(s2_stu.class_id, "?"),
                "resonance": weight,
                "details": edge_details.get((s1, s2), [])
            })

    # ── 8. 高危团伙按共振度降序，取Top10 ──
    gang_warnings.sort(key=lambda x: -x["resonance"])
    gang_warnings = gang_warnings[:10]

    return jsonify({
        "nodes": nodes,
        "links": links,
        "gang_warnings": gang_warnings,
        "total_students_in_graph": len(nodes),
        "total_edges": len(links),
        "data_source": "WingsScore + DisciplineRecord 时空共时性分析",
        "algorithm": "30分钟时间窗口 + 共现引力聚类"
    })


@bigscreen_bp.route("/export/pdf")
@login_required
@require_role("ms_admin", "grade_leader")
def export_pdf():
    """导出PDF报告"""
    from flask import make_response
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import cm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from io import BytesIO
    import os
    from flask import current_app

    # 注册中文字体
    font_path = os.path.join(current_app.root_path, '..', 'static', 'fonts', 'simhei.ttf')
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont('SimHei', font_path))
        font_name = 'SimHei'
    else:
        font_name = 'Helvetica'
    
    # 创建PDF
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # 标题
    c.setFont(font_name, 24)
    c.drawString(2*cm, height - 3*cm, "梨江中学德育数据报告")
    
    # 日期
    c.setFont(font_name, 12)
    now = get_local_now().strftime('%Y-%m-%d %H:%M')
    c.drawString(2*cm, height - 4*cm, f"生成时间: {now}")
    
    # 获取数据
    from models import Student, Class, DisciplineRecord, RoutineScore
    
    # 统计数据
    y = height - 6*cm
    c.setFont(font_name, 14)
    c.drawString(2*cm, y, "一、基本数据")
    y -= 1*cm
    
    c.setFont(font_name, 11)
    total_students = Student.query.filter_by(is_active=True).count()
    total_classes = Class.query.filter_by(is_active=True).count()
    c.drawString(3*cm, y, f"在校学生: {total_students} 人")
    y -= 0.8*cm
    c.drawString(3*cm, y, f"班级数量: {total_classes} 个")
    y -= 0.8*cm
    
    # 违纪统计
    since = get_local_now() - timedelta(days=30)
    disc_count = DisciplineRecord.query.filter(DisciplineRecord.created_at >= since).count()
    c.drawString(3*cm, y, f"近30天违纪: {disc_count} 人次")
    y -= 1.5*cm
    
    c.setFont(font_name, 14)
    c.drawString(2*cm, y, "二、班级常规评分")
    y -= 1*cm
    
    # 班级评分
    c.setFont(font_name, 10)
    from sqlalchemy import func
    class_scores = db.session.query(
        Class.name, func.avg(RoutineScore.score).label("avg_score")
    ).join(RoutineScore, RoutineScore.class_id == Class.id)\
     .filter(Class.is_active == True)\
     .group_by(Class.id).order_by(func.avg(RoutineScore.score).desc()).limit(10).all()
    
    for name, score in class_scores:
        y -= 0.7*cm
        if y < 3*cm:
            c.showPage()
            y = height - 3*cm
        c.drawString(3*cm, y, f"{name}: {round(float(score), 1)} 分")
    
    c.save()
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=deyu_report.pdf'
    return response


@bigscreen_bp.route("/export/excel")
@login_required
@require_role("ms_admin", "grade_leader")
def export_excel():
    """导出Excel报告"""
    from flask import make_response
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from io import BytesIO
    from models import Student, Class, DisciplineRecord, RoutineScore
    from sqlalchemy import func
    from datetime import datetime, timedelta
    
    wb = Workbook()
    
    # ── Sheet1: 基本数据 ──
    ws1 = wb.active
    ws1.title = "基本数据"
    
    row = 1
    ws1['A1'] = "梨江中学德育数据报告"
    ws1['A1'].font = Font(size=16, bold=True)
    row += 1
    
    ws1[f'A{row}'] = f"生成时间: {get_local_now().strftime('%Y-%m-%d %H:%M')}"
    row += 2
    
    ws1[f'A{row}'] = "指标"
    ws1[f'B{row}'] = "数值"
    ws1[f'A{row}'].font = Font(bold=True)
    ws1[f'B{row}'].font = Font(bold=True)
    row += 1
    
    total_students = Student.query.filter_by(is_active=True).count()
    total_classes = Class.query.filter_by(is_active=True).count()
    since = get_local_now() - timedelta(days=30)
    disc_count = DisciplineRecord.query.filter(DisciplineRecord.created_at >= since).count()
    
    ws1[f'A{row}'] = "在校学生"
    ws1[f'B{row}'] = total_students
    row += 1
    
    ws1[f'A{row}'] = "班级数量"
    ws1[f'B{row}'] = total_classes
    row += 1
    
    ws1[f'A{row}'] = "近30天违纪人次"
    ws1[f'B{row}'] = disc_count
    row += 1
    
    # ── Sheet2: 班级评分 ──
    ws2 = wb.create_sheet("班级评分")
    ws2['A1'] = "班级名称"
    ws2['B1'] = "平均评分"
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    
    class_scores = db.session.query(
        Class.name, func.avg(RoutineScore.score).label("avg_score")
    ).join(RoutineScore, RoutineScore.class_id == Class.id)\
     .filter(Class.is_active == True)\
     .group_by(Class.id).order_by(func.avg(RoutineScore.score).desc()).all()
    
    r = 2
    for name, score in class_scores:
        ws2[f'A{r}'] = name
        ws2[f'B{r}'] = round(float(score), 1)
        r += 1
    
    # ── 输出 ──
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = make_response(buffer.read())
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=deyu_report.xlsx'
    return response

"""德育处工作台 — 规则配置/任务下发/问题学生建档/全校总览"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session, current_app
from models import db, Student, Class, Grade, User, Task, TaskFeedback
from models import DisciplineRecord, DisciplineAppeal, RoutineScore, ProblemStudent, ProblemTrack, ROLES, Subject, FlagEvaluation, FlagReport
from models import Attendance, LeaveRequest
from models import Message
from blueprints.discipline_utils import check_escalation, send_discipline_notifications, send_appeal_notifications, deduct_quality_score
from blueprints.audit_log import audit_log
from decorators import login_required, require_role, scope_query, require_permission
from utils.db_utils import safe_commit
from utils import get_local_now
from utils.llm_client import call_llm_json, LLMAvailabilityError
from datetime import date, datetime, timedelta
from sqlalchemy import func
import json as _json

ms_bp = Blueprint("ms", __name__)


# ── 工作台首页 ──
@ms_bp.route("/")
@require_role("ms_admin")
def dashboard():
    stats = {
        "student_count": Student.query.filter_by(is_active=True).count(),
        "class_count": Class.query.filter_by(is_active=True).count(),
        "discipline_count": DisciplineRecord.query.filter_by(status="active").count(),
        "problem_count": ProblemStudent.query.filter_by(status="active").count(),
        "pending_tasks": Task.query.filter_by(status="pending").count(),
    }
    grades = Grade.query.order_by(Grade.sort_order).all()
    recent_disciplines = DisciplineRecord.query.order_by(
        DisciplineRecord.created_at.desc()
    ).limit(10).all()
    return render_template("ms/dashboard.html", stats=stats, grades=grades,
                           recent_disciplines=recent_disciplines)


# ── 任务管理 ──
@ms_bp.route("/tasks")
@require_role("ms_admin")
def task_list():
    status = request.args.get("status", "")
    q = Task.query.filter_by(from_user_id=session.get("user_id"))
    if status:
        q = q.filter_by(status=status)
    tasks = q.order_by(Task.created_at.desc()).all()
    grades = Grade.query.order_by(Grade.sort_order).all()
    classes = Class.query.filter_by(is_active=True).order_by(Class.name).all()
    # 统计各状态数量
    status_counts = {
        "pending": Task.query.filter_by(from_user_id=session.get("user_id"), status="pending").count(),
        "assigned": Task.query.filter_by(from_user_id=session.get("user_id"), status="assigned").count(),
        "done": Task.query.filter_by(from_user_id=session.get("user_id"), status="done").count(),
        "closed": Task.query.filter_by(from_user_id=session.get("user_id"), status="closed").count(),
    }
    return render_template("ms/tasks.html", tasks=tasks, grades=grades, classes=classes,
                           status=status, status_counts=status_counts)


@ms_bp.route("/tasks/create", methods=["POST"])
@require_role("ms_admin")
def create_task():
    target_type = request.form["target_type"]
    # 解析截止日期
    deadline_str = request.form.get("deadline")
    deadline = date.fromisoformat(deadline_str) if deadline_str else None
    if target_type == "grade":
        # 多发：每个选中的年级一条任务
        grade_ids = request.form.getlist("target_ids")
        if not grade_ids:
            flash("请选择目标年级", "danger")
            return redirect(url_for("ms.task_list"))
        for gid in grade_ids:
            task = Task(
                title=request.form["title"],
                content=request.form.get("content", ""),
                from_role="ms_admin",
                from_user_id=session.get("user_id"),
                target_type="grade",
                target_id=int(gid),
                deadline=deadline,
            )
            db.session.add(task)
    elif target_type == "class":
        class_ids = request.form.getlist("target_ids")
        if not class_ids:
            flash("请选择目标班级", "danger")
            return redirect(url_for("ms.task_list"))
        for cid in class_ids:
            task = Task(
                title=request.form["title"],
                content=request.form.get("content", ""),
                from_role="ms_admin",
                from_user_id=session.get("user_id"),
                target_type="class",
                target_id=int(cid),
                deadline=deadline,
            )
            db.session.add(task)
    else:
        flash("无效的目标类型", "danger")
        return redirect(url_for("ms.task_list"))
    safe_commit()
    flash(f"任务已下发", "success")
    return redirect(url_for("ms.task_list"))


@ms_bp.route("/tasks/<int:tid>")
@require_role("ms_admin")
def task_detail(tid):
    task = Task.query.get_or_404(tid)
    feedbacks = (TaskFeedback.query.filter_by(task_id=tid)
                 .order_by(TaskFeedback.created_at.asc()).all())
    # 目标名称
    target_label = "全部"
    if task.target_type == "grade":
        g = Grade.query.get(task.target_id)
        target_label = f"年级：{g.name}" if g else f"年级ID:{task.target_id}"
    elif task.target_type == "class":
        c = Class.query.get(task.target_id)
        target_label = f"班级：{c.name}" if c else f"班级ID:{task.target_id}"
    return render_template("ms/task_detail.html", task=task, feedbacks=feedbacks,
                           target_label=target_label)


@ms_bp.route("/tasks/<int:tid>/close", methods=["POST"])
@require_role("ms_admin")
def close_task(tid):
    task = Task.query.get_or_404(tid)
    task.status = "closed"
    task.finished_at = get_local_now()
    safe_commit()
    flash("任务已关闭", "success")
    return redirect(url_for("ms.task_list"))


# ── 纪律管理 — 全校视图 ──
@ms_bp.route("/discipline")
@require_role("ms_admin")
@require_permission("view_discipline")  # ← 新增：细粒度权限检查
def discipline_list():
    page = request.args.get("page", 1, type=int)
    grade_filter = request.args.get("grade_id", type=int)
    from sqlalchemy.orm import joinedload
    q = DisciplineRecord.query
    if grade_filter:
        q = q.filter_by(grade_id=grade_filter)
    records = q.options(
        joinedload(DisciplineRecord.student).joinedload(Student.class_)
    ).order_by(DisciplineRecord.created_at.desc()).paginate(
        page=page, per_page=20)
    grades = Grade.query.all()
    students = Student.query.filter_by(is_active=True).options(
        joinedload(Student.class_)
    ).order_by(Student.grade_id, Student.class_id, Student.student_no).all()
    return render_template("ms/discipline.html", records=records, grades=grades,
                           students=students, grade_filter=grade_filter)


@ms_bp.route("/discipline/add", methods=["POST"])
@require_role("ms_admin")
@require_permission("manage_discipline")  # ← 新增：细粒度权限检查
def add_discipline():
    from utils.db_utils import safe_commit
    student_id = request.form.get("student_id", type=int)
    student = Student.query.get(student_id)
    if not student:
        flash("学生不存在", "danger")
        return redirect(url_for("ms.discipline_list"))
    record = DisciplineRecord(
        student_id=student.id,
        class_id=student.class_id,
        grade_id=student.grade_id,
        type=request.form["type"],
        category=request.form.get("category", ""),
        description=request.form["description"],
        action_taken=request.form.get("action_taken", ""),
        points=request.form.get("points", 0, type=int),
        created_by=session.get("user_id"),
    )
    db.session.add(record)
    # 积分累计自动升级 + 自动推送通知（与违纪记录同一个事务）
    check_escalation(student, session.get("user_id"))
    send_discipline_notifications(record, student)
    deduct_quality_score(record, student, session.get("user_id"))

    safe_commit()
    flash(f"已记录 {student.name} 违纪", "success")

    # ── MLOps 自进化: 违纪记录变更触发模型自动重训 ──
    try:
        from utils.model_retrain import trigger_auto_retrain
        trigger_auto_retrain(current_app._get_current_object(), grade_id=student.grade_id)
    except Exception:
        pass  # 重训失败不影响违纪记录

    return redirect(url_for("ms.discipline_list"))


@ms_bp.route("/discipline/<int:rid>/delete", methods=["POST"])
@require_role("ms_admin")
def delete_discipline(rid):
    record = DisciplineRecord.query.get_or_404(rid)
    db.session.delete(record)
    safe_commit()
    flash("违纪记录已删除", "success")
    return redirect(url_for("ms.discipline_list"))


@ms_bp.route("/discipline/stats")
@require_role("ms_admin")
def discipline_stats():
    import json
    grades = Grade.query.all()
    # 按年级统计 — 单次 GROUP BY 替代 N×3 次 COUNT
    grade_raw = db.session.query(
        DisciplineRecord.grade_id,
        func.count().label("total"),
        func.sum(func.if_(DisciplineRecord.status == "active", 1, 0)).label("active"),
        func.sum(func.if_(DisciplineRecord.status == "resolved", 1, 0)).label("resolved"),
    ).group_by(DisciplineRecord.grade_id).all()
    grade_map = {}
    for row in grade_raw:
        grade_map[row[0]] = {"total": int(row[1]), "active": int(row[2] or 0), "resolved": int(row[3] or 0)}
    grade_stats = []
    for g in grades:
        s = grade_map.get(g.id, {"total": 0, "active": 0, "resolved": 0})
        grade_stats.append({"name": g.name, "total": s["total"], "active": s["active"], "resolved": s["resolved"]})

    # 按类型统计 + 中文标签
    type_labels_map = {"warning": "警告", "minor": "轻微", "major": "重大", "serious": "严重"}
    type_stats = db.session.query(
        DisciplineRecord.type, func.count().label("cnt")
    ).group_by(DisciplineRecord.type).all()
    type_stats = [{"type": type_labels_map.get(t[0], t[0]), "cnt": t[1]} for t in type_stats]

    # 按类别统计
    category_stats = db.session.query(
        DisciplineRecord.category,
        func.count().label("cnt"),
        func.sum(DisciplineRecord.points).label("points_sum")
    ).group_by(DisciplineRecord.category).all()
    category_stats = [{"category": c[0] or "未分类", "cnt": int(c[1]), "points_sum": float(c[2]) if c[2] else 0} for c in category_stats]

    # 按班级统计（柱状图数据）
    class_stats = db.session.query(
        Class.name, func.count().label("cnt")
    ).join(DisciplineRecord, DisciplineRecord.class_id == Class.id).group_by(
        DisciplineRecord.class_id, Class.name
    ).order_by(Class.grade_id, Class.name).all()
    class_stats = [{"name": c[0], "cnt": c[1]} for c in class_stats]

    # 转为JSON供Chart.js使用
    chart_data = {
        "typeLabels": json.dumps([t["type"] for t in type_stats]),
        "typeCounts": json.dumps([t["cnt"] for t in type_stats]),
        "categoryLabels": json.dumps([c["category"] for c in category_stats]),
        "categoryCounts": json.dumps([c["cnt"] for c in category_stats]),
        "categoryPoints": json.dumps([c["points_sum"] for c in category_stats]),
        "classLabels": json.dumps([c["name"] for c in class_stats]),
        "classCounts": json.dumps([c["cnt"] for c in class_stats]),
    }

    return render_template("ms/discipline_stats.html",
                           grade_stats=grade_stats,
                           type_stats=type_stats,
                           category_stats=category_stats,
                           class_stats=class_stats,
                           chart_data=chart_data)


@ms_bp.route("/discipline/<int:rid>/resolve", methods=["POST"])
@require_role("ms_admin")
def discipline_resolve(rid):
    record = DisciplineRecord.query.get_or_404(rid)
    record.status = "resolved"
    record.resolved_at = get_local_now()
    safe_commit()
    flash("已标记为已解决", "success")
    return redirect(url_for("ms.discipline_list"))


# ── 纪律申诉复核 ──────────────────────────────────────────────────
@ms_bp.route("/appeals")
@require_role("ms_admin")
def appeal_list():
    """申诉列表"""
    status_filter = request.args.get("status", "")
    page = request.args.get("page", 1, type=int)

    q = DisciplineAppeal.query
    if status_filter:
        q = q.filter_by(status=status_filter)
    q = q.order_by(DisciplineAppeal.status == "pending",
                   DisciplineAppeal.created_at.desc() if status_filter != "pending" else DisciplineAppeal.created_at.asc())

    pagination = q.paginate(page=page, per_page=20, error_out=False)
    appeals = pagination.items

    # 统计数据
    pending_count = DisciplineAppeal.query.filter_by(status="pending").count()
    approved_count = DisciplineAppeal.query.filter_by(status="approved").count()
    rejected_count = DisciplineAppeal.query.filter_by(status="rejected").count()

    return render_template("ms/appeals.html",
                           appeals=appeals,
                           pagination=pagination,
                           status_filter=status_filter,
                           pending_count=pending_count,
                           approved_count=approved_count,
                           rejected_count=rejected_count)


@ms_bp.route("/appeals/<int:aid>", methods=["GET", "POST"])
@require_role("ms_admin")
def appeal_review(aid):
    """申诉复核"""
    appeal = DisciplineAppeal.query.get_or_404(aid)

    if request.method == "POST":
        action = request.form.get("action", "")
        review_comment = request.form.get("review_comment", "").strip()

        if action == "approve":
            appeal.status = "approved"
            # 撤销原始违纪记录
            record = appeal.discipline
            if record:
                record.status = "resolved"
                record.resolved_at = get_local_now()
            flash("申诉已通过，原违纪记录已撤销", "success")
        elif action == "reject":
            appeal.status = "rejected"
            # 恢复违纪记录状态
            record = appeal.discipline
            if record and record.status == "appealed":
                record.status = "active"
            flash("申诉已驳回", "warning")
        else:
            flash("无效操作", "danger")
            return redirect(url_for("ms.appeal_review", aid=aid))

        appeal.review_comment = review_comment or ("同意申诉" if action == "approve" else "申诉理由不成立")
        appeal.reviewed_by = session.get("user_id")
        appeal.reviewed_at = get_local_now()
        safe_commit()

        send_appeal_notifications(appeal, appeal.student, appeal.discipline)
        safe_commit()

        return redirect(url_for("ms.appeal_list"))

    # 获取该学生的违纪记录历史
    discipline_history = DisciplineRecord.query.filter_by(
        student_id=appeal.student_id
    ).order_by(DisciplineRecord.created_at.desc()).limit(10).all()

    return render_template("ms/appeal_review.html",
                           appeal=appeal,
                           discipline_history=discipline_history)


# ── 常规评分总览 ──
@ms_bp.route("/routine")
@require_role("ms_admin")
def routine_overview():
    grade_id = request.args.get("grade_id", type=int)
    page = request.args.get("page", 1, type=int)
    category = request.args.get("category", type=str)
    per_page = 50

    scores = RoutineScore.query
    if grade_id:
        scores = scores.filter_by(grade_id=grade_id)
    if category:
        scores = scores.filter_by(category=category)
    scores = scores.order_by(RoutineScore.record_date.desc()).paginate(page=page, per_page=per_page, error_out=False)
    grades = Grade.query.all()
    classes = Class.query.filter_by(is_active=True).all()
    classes_by_id = {c.id: c for c in classes}
    return render_template("ms/routine.html", scores=scores, grades=grades,
                           classes=classes, classes_by_id=classes_by_id,
                           today=date.today(), grade_filter=grade_id,
                           category_filter=category)


@ms_bp.route("/routine/add", methods=["POST"])
@require_role("ms_admin")
def add_routine():
    class_id = request.form.get("class_id", type=int)
    cls = Class.query.get(class_id)
    if not cls:
        flash("班级不存在", "danger")
        return redirect(url_for("ms.routine_overview"))
    record_date = request.form.get("record_date", str(date.today()))
    score = RoutineScore(
        class_id=class_id,
        grade_id=cls.grade_id,
        category=request.form["category"],
        score=request.form.get("score", 0, type=int),
        note=request.form.get("note", ""),
        inspector=request.form.get("inspector", session.get("display_name", "")),
        scorer_type="ms_admin",
        record_date=date.fromisoformat(record_date) if record_date else date.today(),
    )
    db.session.add(score)
    safe_commit()
    flash("评分已录入", "success")
    return redirect(url_for("ms.routine_overview"))


@ms_bp.route("/routine/<int:sid>/delete", methods=["POST"])
@require_role("ms_admin")
def delete_routine(sid):
    s = RoutineScore.query.get_or_404(sid)
    db.session.delete(s)
    safe_commit()
    flash("评分已删除", "success")
    return redirect(url_for("ms.routine_overview"))


# ── 班级流动红旗（三维度加权评价）──

def _calc_flag_weights(self_score, grade_score, ms_score):
    """
    根据三维度数据是否可用，计算实际权重。
    标准权重: 班主任0.2 + 年级组0.3 + 德育处0.5
    某维度缺失时，其余维度按比例瓜分该权重。
    """
    BASE_W = [0.2, 0.3, 0.5]
    scores = [float(s) if s is not None else None for s in (self_score, grade_score, ms_score)]
    available = [s is not None for s in scores]

    if not any(available):
        return BASE_W, 0.0

    if all(available):
        return BASE_W, (self_score * 0.2 + grade_score * 0.3 + ms_score * 0.5)

    # 缺失维度权重按比例重分配
    missing_weight = sum(w for w, a in zip(BASE_W, available) if not a)
    avail_indices = [i for i, a in enumerate(available) if a]
    avail_total_w = sum(BASE_W[i] for i in avail_indices)

    weights = list(BASE_W)
    for i in avail_indices:
        weights[i] = BASE_W[i] + missing_weight * (BASE_W[i] / avail_total_w)
    for i in range(3):
        if not available[i]:
            weights[i] = 0.0

    final = sum(s * w for s, w in zip(scores, weights) if s is not None)
    return weights, round(final, 2)


@ms_bp.route("/leaderboard")
@require_role("ms_admin")
def leaderboard():
    """流动红旗评价 — 支持 Tab 切换周评/月评/期末"""
    grade_id = request.args.get("grade_id", type=int)
    period_type = request.args.get("period_type", "week")
    period_label = request.args.get("period_label", "")

    grades = Grade.query.all()

    # 计算可选的期间列表
    now = get_local_now()
    period_labels = _build_period_labels(now)

    # 如果没选期间，默认选第一个
    if not period_label and period_labels.get(period_type):
        period_label = period_labels[period_type][0]["value"]

    # 查询已发布的评价记录
    q = FlagEvaluation.query.filter_by(
        period_type=period_type,
        status="published"
    )
    if grade_id:
        q = q.filter_by(grade_id=grade_id)
    if period_label:
        q = q.filter_by(period_label=period_label)

    evals = q.order_by(FlagEvaluation.rank.asc()).all()
    # 按年级分组
    evals_by_grade = {}
    for ev in evals:
        evals_by_grade.setdefault(ev.grade_id, []).append(ev)

    return render_template("ms/leaderboard.html",
                           grades=grades, grade_filter=grade_id,
                           period_type=period_type, period_label=period_label,
                           period_labels=period_labels,
                           evals_by_grade=evals_by_grade)


def _build_period_labels(now):
    """生成当前可用的周/月/期末期间列表"""
    import calendar

    # 周列表：最近8周（按周一）
    week_labels = []
    # 找到本周一
    weekday = now.weekday()
    this_monday = now.date() - timedelta(days=weekday)
    for i in range(7, -1, -1):
        mon = this_monday - timedelta(weeks=i)
        sun = mon + timedelta(days=6)
        week_num = mon.isocalendar()[1]
        week_labels.append({
            "value": mon.isoformat(),
            "label": f"第{week_num}周 ({mon.strftime('%m/%d')}~{sun.strftime('%m/%d')})",
            "start": mon,
            "end": sun,
        })

    # 月列表：最近6个月（纯标准库实现）
    month_labels = []
    for i in range(5, -1, -1):
        # 用整数月份算术代替 dateutil
        y, m = now.year, now.month
        m = m - i
        while m <= 0:
            m += 12
            y -= 1
        d = date(y, m, 1)
        _, last_day = calendar.monthrange(d.year, d.month)
        month_labels.append({
            "value": d.isoformat(),
            "label": f"{d.year}年{d.month}月",
            "start": d,
            "end": date(d.year, d.month, last_day),
        })

    # 期末标签
    term_labels = [{"value": "2025-2026-2", "label": "2025-2026学年第二学期"}]

    return {
        "week": week_labels,
        "month": month_labels,
        "term": term_labels,
    }


@ms_bp.route("/leaderboard/generate", methods=["POST"])
@require_role("ms_admin")
def generate_evaluation():
    """生成评价草稿 — 根据选定期间和年级，计算三维度加权得分"""
    grade_id = request.form.get("grade_id", type=int)
    period_type = request.form.get("period_type")
    period_label = request.form.get("period_label")

    if not period_type or not period_label:
        flash("请选择评价周期", "danger")
        return redirect(url_for("ms.leaderboard"))

    # 解析期间日期范围
    start_date, end_date = _parse_period_range(period_type, period_label)
    if not start_date or not end_date:
        flash("无法解析评价周期", "danger")
        return redirect(url_for("ms.leaderboard"))

    # 获取目标班级列表
    if grade_id:
        classes = Class.query.filter_by(grade_id=grade_id, is_active=True).all()
        grades = Grade.query.filter_by(id=grade_id).all()
    else:
        classes = Class.query.filter_by(is_active=True).all()
        grades = Grade.query.all()

    if not classes:
        flash("没有找到班级", "danger")
        return redirect(url_for("ms.leaderboard"))

    # 批量查询该期间内所有常规评分（按 scorer_type 分组）
    all_scores = RoutineScore.query.filter(
        RoutineScore.record_date >= start_date,
        RoutineScore.record_date <= end_date,
        RoutineScore.class_id.in_([c.id for c in classes])
    ).all()

    # 按班级+scorer_type聚合均分
    from collections import defaultdict
    agg = defaultdict(lambda: defaultdict(list))
    for s in all_scores:
        agg[s.class_id][s.scorer_type].append(s.score)

    # 删除该期间的旧草稿
    FlagEvaluation.query.filter_by(
        period_type=period_type, period_label=period_label, status="draft"
    ).filter(
        FlagEvaluation.class_id.in_([c.id for c in classes])
    ).delete(synchronize_session="fetch")
    safe_commit()

    # 生成新草稿
    created = 0
    for cls in classes:
        self_scores = agg.get(cls.id, {}).get("class_teacher", [])
        grade_scores = agg.get(cls.id, {}).get("grade_leader", [])
        ms_scores = agg.get(cls.id, {}).get("ms_admin", [])

        self_avg = sum(self_scores) / len(self_scores) if self_scores else None
        grade_avg = sum(grade_scores) / len(grade_scores) if grade_scores else None
        ms_avg = sum(ms_scores) / len(ms_scores) if ms_scores else None

        weights, base = _calc_flag_weights(self_avg, grade_avg, ms_avg)

        # ── 违纪+考勤合流扣分 ──
        # 1. 违纪扣分（周期内该班级所有违纪记录的points总和 × 0.1）
        discipline_points = float(
            db.session.query(func.sum(DisciplineRecord.points))
            .filter(
                DisciplineRecord.class_id == cls.id,
                DisciplineRecord.created_at >= start_date,
                DisciplineRecord.created_at <= end_date,
            ).scalar() or 0
        )

        # 2. 考勤异常扣分（迟到/缺勤/早退/请假 次数 × 0.05）
        attendance_exceptions = Attendance.query.filter(
            Attendance.class_id == cls.id,
            Attendance.status.in_(["late", "absent", "early", "leave"]),
            Attendance.record_date >= start_date,
            Attendance.record_date <= end_date
        ).count()

        # 扣分系数按周期类型动态调整
        # week: 0.1  month: 0.05  term: 0.01
        _discipline_coeff = {"week": 0.1, "month": 0.05, "term": 0.01}.get(period_type, 0.1)
        _attendance_coeff = {"week": 0.05, "month": 0.03, "term": 0.01}.get(period_type, 0.05)

        discipline_deduction = round(discipline_points * _discipline_coeff, 2)
        attendance_deduction = round(attendance_exceptions * _attendance_coeff, 2)

        final = round(base - discipline_deduction - attendance_deduction, 2)
        final = max(0.0, final)  # 防止扣成负数

        ev = FlagEvaluation(
            period_type=period_type,
            period_label=period_label,
            grade_id=cls.grade_id,
            class_id=cls.id,
            self_score=self_avg,
            grade_score=grade_avg,
            ms_score=ms_avg,
            self_weight=weights[0],
            grade_weight=weights[1],
            ms_weight=weights[2],
            base_score=base,
            discipline_points=discipline_points,
            discipline_deduction=discipline_deduction,
            attendance_exceptions=attendance_exceptions,
            attendance_deduction=attendance_deduction,
            final_score=final,
            status="draft",
        )
        db.session.add(ev)
        created += 1

    safe_commit()
    flash(f"已生成 {created} 个班级的评价草稿，请审核后发布", "success")

    return redirect(url_for("ms.leaderboard", period_type=period_type, period_label=period_label,
                             grade_id=grade_id))


@ms_bp.route("/leaderboard/publish", methods=["POST"])
@require_role("ms_admin")
def publish_evaluation():
    """发布评价 — 计算排名并标记为已发布"""
    grade_id = request.form.get("grade_id", type=int)
    period_type = request.form.get("period_type")
    period_label = request.form.get("period_label")

    if not period_type or not period_label:
        flash("请选择评价周期", "danger")
        return redirect(url_for("ms.leaderboard"))

    # 查找草稿
    q = FlagEvaluation.query.filter_by(
        period_type=period_type, period_label=period_label, status="draft"
    )
    if grade_id:
        q = q.filter_by(grade_id=grade_id)

    drafts = q.all()
    if not drafts:
        flash("没有找到待发布的草稿", "danger")
        return redirect(url_for("ms.leaderboard"))

    # 按年级分组计算排名
    from itertools import groupby
    drafts.sort(key=lambda x: (x.grade_id, -x.final_score))
    for g_id, group in groupby(drafts, key=lambda x: x.grade_id):
        group_list = list(group)
        for rank, ev in enumerate(group_list, 1):
            ev.rank = rank
            ev.status = "published"
            ev.published_at = get_local_now()

    safe_commit()
    flash(f"已发布 {len(drafts)} 个班级的评价结果", "success")
    return redirect(url_for("ms.leaderboard", period_type=period_type, period_label=period_label,
                             grade_id=grade_id))


@ms_bp.route("/leaderboard/drafts")
@require_role("ms_admin")
def view_drafts():
    """查看待发布的草稿"""
    grade_id = request.args.get("grade_id", type=int)
    period_type = request.args.get("period_type", "week")
    period_label = request.args.get("period_label", "")

    grades = Grade.query.all()
    period_labels = _build_period_labels(get_local_now())

    if not period_label and period_labels.get(period_type):
        period_label = period_labels[period_type][0]["value"]

    q = FlagEvaluation.query.filter_by(
        period_type=period_type, period_label=period_label, status="draft"
    )
    if grade_id:
        q = q.filter_by(grade_id=grade_id)

    drafts = q.order_by(FlagEvaluation.final_score.desc()).all()
    evals_by_grade = {}
    for ev in drafts:
        evals_by_grade.setdefault(ev.grade_id, []).append(ev)

    return render_template("ms/leaderboard_drafts.html",
                           grades=grades, grade_filter=grade_id,
                           period_type=period_type, period_label=period_label,
                           period_labels=period_labels,
                           evals_by_grade=evals_by_grade)


@ms_bp.route("/leaderboard/api/periods")
@require_role("ms_admin")
def api_periods():
    """返回可选期间列表（AJAX用）"""
    return jsonify(_build_period_labels(get_local_now()))


def _parse_period_range(period_type, period_label):
    """解析期间标签为日期范围 (start_date, end_date)"""
    import calendar
    try:
        if period_type == "term":
            # 期末评价覆盖整个学期，返回一个极宽范围
            return date(2026, 2, 16), date(2026, 7, 10)

        elif period_type == "week":
            # period_label 是周一的 ISO 日期
            start = date.fromisoformat(period_label)
            end = start + timedelta(days=6)
            return start, end

        elif period_type == "month":
            # period_label 是月初的 ISO 日期
            start = date.fromisoformat(period_label)
            import calendar
            _, last_day = calendar.monthrange(start.year, start.month)
            end = date(start.year, start.month, last_day)
            return start, end
    except (ValueError, TypeError):
        pass
    return None, None


# ══════════════════════════════════════════════════════════════
#  流动红旗周报 — AI 一键生成
# ══════════════════════════════════════════════════════════════

FLAG_REPORT_SYSTEM_PROMPT = """你是一位拥有20年经验的初中德育管理资深主任，擅长撰写流动红旗评价周报。你的报告将直接用于全校德育工作通报，语气要有权威感又有温度。

[核心约束]
1. 严禁出现"在领导的关怀下"等空话套话。
2. 必须引用具体数据：班级得分、排名变动、违纪扣分详情、考勤异常次数。
3. 颁奖词要有仪式感和文学性，末位建议要体现关怀而非批评。
4. report_summary 200-400字，award_speech 100-150字，runner_up_speech 80-120字，improvement_advice 100-200字。
5. highlights 是3-5个关键数据洞察（如"XX班考勤异常达N次，为主要失分项"）。

[输出格式]
必须输出严格的 JSON 格式，不要任何 Markdown 包裹（如 ```json），直接返回以下结构的字符串：
{
  "report_summary": "200-400字周报总述，涵盖整体表现、排名格局、关键变化",
  "award_speech": "100-150字颁奖词，授予第一名班级，要有仪式感",
  "runner_up_speech": "80-120字亚军寄语，鼓励继续追赶",
  "improvement_advice": "100-200字改进建议，针对末位班级，语气温暖有建设性",
  "highlights": ["关键洞察1", "关键洞察2", "关键洞察3"]
}"""


def _build_flag_report_context(evaluations):
    """
    根据 FlagEvaluation 已发布列表，构建 LLM 上下文字符串。
    evaluations: 同一 (period_type, period_label, grade_id) 下的已发布评价列表（按 rank 排序）
    """
    lines = []
    lines.append(f"评价类型: {evaluations[0].period_type}")
    lines.append(f"评价周期: {evaluations[0].period_label}")
    lines.append(f"参评班级数: {len(evaluations)}")
    lines.append("")

    for ev in evaluations:
        cls_name = ev.class_.name if ev.class_ else "未知班级"
        rank_str = f"第{ev.rank}名" if ev.rank else "未排名"
        score_parts = []
        if ev.self_score is not None:
            score_parts.append(f"班主任自评={ev.self_score:.1f}(权重{ev.self_weight:.0%})")
        if ev.grade_score is not None:
            score_parts.append(f"年级组评级={ev.grade_score:.1f}(权重{ev.grade_weight:.0%})")
        if ev.ms_score is not None:
            score_parts.append(f"德育处评级={ev.ms_score:.1f}(权重{ev.ms_weight:.0%})")

        line = f"{rank_str} {cls_name}: 最终得分 {ev.final_score:.1f}"
        if score_parts:
            line += f" | 维度分解: {', '.join(score_parts)}"
        if ev.base_score is not None:
            line += f" | 加权底分 {ev.base_score:.1f}"
        if ev.discipline_deduction and ev.discipline_deduction > 0:
            line += f" | 违纪扣分 -{ev.discipline_deduction:.1f} (总违纪{ev.discipline_points:.0f}分)"
        if ev.attendance_deduction and ev.attendance_deduction > 0:
            line += f" | 考勤扣分 -{ev.attendance_deduction:.1f} (异常{ev.attendance_exceptions}次)"
        lines.append(line)

    return "\n".join(lines)


@ms_bp.route("/flag-reports")
@require_role("ms_admin")
def flag_report_list():
    """流动红旗周报列表 — 按年级分组展示"""
    grade_id = request.args.get("grade_id", type=int)
    grades = Grade.query.all()

    q = FlagReport.query.order_by(FlagReport.created_at.desc())
    if grade_id:
        q = q.filter_by(grade_id=grade_id)

    reports = q.all()

    # 按年级分组
    from itertools import groupby
    reports.sort(key=lambda r: r.grade_id)
    reports_by_grade = {}
    for g_id, group in groupby(reports, key=lambda r: r.grade_id):
        reports_by_grade[g_id] = list(group)

    return render_template("ms/flag_report_list.html",
                           reports_by_grade=reports_by_grade,
                           grades=grades,
                           grade_filter=grade_id)


@ms_bp.route("/flag-reports/generate", methods=["POST"])
@require_role("ms_admin")
def generate_flag_report():
    """调用 LLM 生成流动红旗周报"""
    grade_id = request.form.get("grade_id", type=int)
    period_type = request.form.get("period_type")
    period_label = request.form.get("period_label")

    if not period_type or not period_label:
        flash("请选择评价周期", "danger")
        return redirect(url_for("ms.flag_report_list"))

    # 查询该周期+年级的已发布评价
    q = FlagEvaluation.query.filter_by(
        period_type=period_type, period_label=period_label, status="published"
    )
    if grade_id:
        q = q.filter_by(grade_id=grade_id)
    else:
        # 未选年级 → 按年级逐个生成
        grades = Grade.query.all()
        generated = 0
        for g in grades:
            g_evals = q.filter_by(grade_id=g.id).order_by(FlagEvaluation.rank.asc()).all()
            if len(g_evals) < 2:
                continue
            report = _create_flag_report_for_grade(g.id, period_type, period_label, g_evals)
            if report:
                generated += 1

        if generated > 0:
            flash(f"已生成 {generated} 份周报（按年级分组）", "success")
        else:
            flash("该周期没有足够的已发布评价数据（每个年级至少需要2个班级）", "warning")
        return redirect(url_for("ms.flag_report_list"))

    # 选了年级 → 单年级生成
    evals = q.order_by(FlagEvaluation.rank.asc()).all()
    if len(evals) < 2:
        flash("该周期该年级至少需要2个班级的已发布评价才能生成周报", "warning")
        return redirect(url_for("ms.flag_report_list"))

    report = _create_flag_report_for_grade(grade_id, period_type, period_label, evals)
    if report:
        flash("周报生成成功", "success")
        return redirect(url_for("ms.flag_report_detail", rid=report.id))
    else:
        flash("周报生成失败", "danger")
        return redirect(url_for("ms.flag_report_list"))


def _create_flag_report_for_grade(grade_id, period_type, period_label, evals):
    """为单个年级创建流动红旗周报（内部函数）"""
    # 检查是否已存在
    existing = FlagReport.query.filter_by(
        period_type=period_type, period_label=period_label, grade_id=grade_id
    ).first()
    if existing:
        # 覆盖旧报告
        db.session.delete(existing)
        safe_commit()

    # 构建 LLM 上下文
    context = _build_flag_report_context(evals)

    try:
        data = call_llm_json(FLAG_REPORT_SYSTEM_PROMPT, context, max_tokens=2048, timeout=45)
    except Exception as e:
        flash(f"AI 生成失败: {str(e)}", "danger")
        return None

    top_ev = evals[0]  # 已按 rank 排序
    bottom_ev = evals[-1]

    report = FlagReport(
        period_type=period_type,
        period_label=period_label,
        grade_id=grade_id,
        report_summary=data.get("report_summary", ""),
        award_speech=data.get("award_speech", ""),
        runner_up_speech=data.get("runner_up_speech", ""),
        improvement_advice=data.get("improvement_advice", ""),
        highlights=_json.dumps(data.get("highlights", []), ensure_ascii=False),
        class_count=len(evals),
        top_class_id=top_ev.class_id,
        top_class_name=top_ev.class_.name if top_ev.class_ else "",
        top_score=top_ev.final_score,
        bottom_class_id=bottom_ev.class_id,
        bottom_class_name=bottom_ev.class_.name if bottom_ev.class_ else "",
        status="draft",
        created_by=session.get("username", ""),
    )
    db.session.add(report)
    safe_commit()

    audit_log("generate_flag_report", f"生成流动红旗周报: {period_type}/{period_label}/年级{grade_id}")
    return report


@ms_bp.route("/flag-reports/<int:rid>")
@require_role("ms_admin")
def flag_report_detail(rid):
    """流动红旗周报详情"""
    report = FlagReport.query.get_or_404(rid)
    # 解析 highlights JSON
    try:
        highlights = _json.loads(report.highlights) if report.highlights else []
    except Exception:
        highlights = []
    return render_template("ms/flag_report_detail.html", report=report, highlights=highlights)


@ms_bp.route("/flag-reports/<int:rid>/publish", methods=["POST"])
@require_role("ms_admin")
def publish_flag_report(rid):
    """发布周报"""
    report = FlagReport.query.get_or_404(rid)
    report.status = "published"
    safe_commit()
    audit_log("publish_flag_report", f"发布流动红旗周报 ID={rid}")
    flash("周报已发布", "success")
    return redirect(url_for("ms.flag_report_detail", rid=rid))


@ms_bp.route("/flag-reports/<int:rid>/delete", methods=["POST"])
@require_role("ms_admin")
def delete_flag_report(rid):
    """删除周报"""
    report = FlagReport.query.get_or_404(rid)
    db.session.delete(report)
    safe_commit()
    audit_log("delete_flag_report", f"删除流动红旗周报 ID={rid}")
    flash("周报已删除", "success")
    return redirect(url_for("ms.flag_report_list"))


# ── 问题学生管理 — 全校建档 ──
@ms_bp.route("/problem-students")
@require_role("ms_admin")
def problem_list():
    page = request.args.get("page", 1, type=int)
    grade_filter = request.args.get("grade_id", type=int)
    level_filter = request.args.get("level", "")
    q = ProblemStudent.query
    if grade_filter:
        q = q.filter_by(grade_id=grade_filter)
    if level_filter:
        q = q.filter_by(level=level_filter)
    records = q.order_by(ProblemStudent.updated_at.desc()).paginate(page=page, per_page=20)
    grades = Grade.query.all()
    return render_template("ms/problem_list.html", records=records, grades=grades)


@ms_bp.route("/problem-students/create", methods=["GET", "POST"])
@require_role("ms_admin")
def create_problem():
    if request.method == "POST":
        student = Student.query.get(request.form["student_id"])
        if not student:
            flash("学生不存在", "danger")
            return redirect(url_for("ms.create_problem"))
        ps = ProblemStudent(
            student_id=student.id,
            class_id=student.class_id,
            grade_id=student.grade_id,
            category=request.form["category"],
            level=request.form["level"],
            description=request.form["description"],
            intervention=request.form.get("intervention", ""),
            created_by=session.get("user_id"),
        )
        db.session.add(ps)
        safe_commit()
        flash("问题学生已建档", "success")
        return redirect(url_for("ms.problem_list"))
    classes = Class.query.all()
    return render_template("ms/problem_create.html", classes=classes)


@ms_bp.route("/problem-students/<int:pid>")
@require_role("ms_admin")
def problem_detail(pid):
    ps = ProblemStudent.query.get_or_404(pid)
    tracks = ProblemTrack.query.filter_by(problem_id=pid).order_by(
        ProblemTrack.created_at.desc()).all()
    return render_template("ms/problem_detail.html", problem=ps, tracks=tracks)


@ms_bp.route("/problem-students/<int:pid>/track", methods=["POST"])
@require_role("ms_admin")
def add_track(pid):
    track = ProblemTrack(
        problem_id=pid,
        content=request.form["content"],
        created_by=session.get("user_id"),
    )
    db.session.add(track)
    safe_commit()
    flash("跟踪记录已添加", "success")
    return redirect(url_for("ms.problem_detail", pid=pid))


# ── 学生搜索 API ──
@ms_bp.route("/api/students/search")
@require_role("ms_admin", "grade_leader", "class_teacher")
def search_students():
    q = request.args.get("q", "")
    class_id = request.args.get("class_id", type=int)
    query = Student.query.filter(Student.name.contains(q))
    if class_id:
        query = query.filter_by(class_id=class_id)
    students = query.limit(20).all()
    return jsonify([{"id": s.id, "name": s.name, "class": s.class_.name if s.class_ else ""}
                    for s in students])


# ────────────────────────────────────────────
# 年级管理
# ────────────────────────────────────────────
@ms_bp.route("/grades")
@require_role("ms_admin")
def grade_manage():
    grades = Grade.query.order_by(Grade.sort_order).all()
    return render_template("ms/grade_manage.html", grades=grades)


@ms_bp.route("/grades/create", methods=["GET", "POST"])
@require_role("ms_admin")
def grade_create():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        sort_order = request.form.get("sort_order", 0, type=int)
        if not name:
            flash("年级名称不能为空", "warning")
            return redirect(url_for("ms.grade_manage"))
        if Grade.query.filter_by(name=name).first():
            flash("该年级已存在", "warning")
            return redirect(url_for("ms.grade_manage"))
        g = Grade(name=name, sort_order=sort_order)
        db.session.add(g)
        safe_commit()
        flash("年级已创建", "success")
        return redirect(url_for("ms.grade_manage"))
    return render_template("ms/grade_form.html", grade=None)


@ms_bp.route("/grades/<int:gid>/edit", methods=["GET", "POST"])
@require_role("ms_admin")
def grade_edit(gid):
    g = Grade.query.get_or_404(gid)
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        sort_order = request.form.get("sort_order", 0, type=int)
        if not name:
            flash("年级名称不能为空", "warning")
            return redirect(url_for("ms.grade_edit", gid=gid))
        exist = Grade.query.filter(Grade.name == name, Grade.id != gid).first()
        if exist:
            flash("该年级名称已被使用", "warning")
            return redirect(url_for("ms.grade_edit", gid=gid))
        g.name = name
        g.sort_order = sort_order
        safe_commit()
        flash("年级已更新", "success")
        return redirect(url_for("ms.grade_manage"))
    return render_template("ms/grade_form.html", grade=g)


@ms_bp.route("/grades/<int:gid>/delete", methods=["POST"])
@require_role("ms_admin")
def grade_delete(gid):
    g = Grade.query.get_or_404(gid)
    if g.classes.filter_by(is_active=True).count() > 0:
        flash("该年级下还有活跃班级，请先删除或停用班级", "danger")
        return redirect(url_for("ms.grade_manage"))
    g.is_active = False
    safe_commit()
    flash("年级已标记删除", "success")
    return redirect(url_for("ms.grade_manage"))


# ────────────────────────────────────────────
# 班级管理
# ────────────────────────────────────────────
@ms_bp.route("/classes")
@require_role("ms_admin")
def class_manage():
    grades = Grade.query.filter_by(is_active=True).order_by(Grade.sort_order).all()
    classes = Class.query.filter_by(is_active=True).order_by(Class.grade_id, Class.name).all()
    # 批量预加载学生人数 — 单次 GROUP BY 替代 N 次 COUNT
    cnt_map = dict(db.session.query(
        Student.class_id, func.count(Student.id)
    ).filter(
        Student.is_active == True,
        Student.class_id.in_([c.id for c in classes])
    ).group_by(Student.class_id).all())
    for c in classes:
        c.student_cnt = cnt_map.get(c.id, 0)
    return render_template("ms/class_manage.html", grades=grades, classes=classes)


@ms_bp.route("/classes/create", methods=["GET", "POST"])
@require_role("ms_admin")
def class_create():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        grade_id = request.form.get("grade_id", type=int)
        if not name or not grade_id:
            flash("班级名称和所属年级不能为空", "warning")
            return redirect(url_for("ms.class_manage"))
        if Class.query.filter_by(name=name, grade_id=grade_id).first():
            flash("该年级下已有同名班级", "warning")
            return redirect(url_for("ms.class_manage"))
        c = Class(name=name, grade_id=grade_id)
        db.session.add(c)
        safe_commit()
        flash("班级已创建", "success")
        return redirect(url_for("ms.class_manage"))
    grades = Grade.query.filter_by(is_active=True).order_by(Grade.sort_order).all()
    return render_template("ms/class_form.html", class_=None, grades=grades)


@ms_bp.route("/classes/<int:cid>/edit", methods=["GET", "POST"])
@require_role("ms_admin")
def class_edit(cid):
    c = Class.query.get_or_404(cid)
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        grade_id = request.form.get("grade_id", type=int)
        if not name or not grade_id:
            flash("班级名称和所属年级不能为空", "warning")
            return redirect(url_for("ms.class_edit", cid=cid))
        exist = Class.query.filter(
            Class.name == name, Class.grade_id == grade_id, Class.id != cid
        ).first()
        if exist:
            flash("该年级下已有同名班级", "warning")
            return redirect(url_for("ms.class_edit", cid=cid))
        c.name = name
        c.grade_id = grade_id
        safe_commit()
        flash("班级已更新", "success")
        return redirect(url_for("ms.class_manage"))
    grades = Grade.query.filter_by(is_active=True).order_by(Grade.sort_order).all()
    return render_template("ms/class_form.html", class_=c, grades=grades)


@ms_bp.route("/classes/<int:cid>/delete", methods=["POST"])
@require_role("ms_admin")
def class_delete(cid):
    c = Class.query.get_or_404(cid)
    cnt = Student.query.filter_by(class_id=cid, is_active=True).count()
    if cnt > 0:
        flash("该班级还有 %d 名活跃学生，请先转移或删除学生" % cnt, "danger")
        return redirect(url_for("ms.class_manage"))
    c.is_active = False
    safe_commit()
    flash("班级已标记删除", "success")
    return redirect(url_for("ms.class_manage"))


# ────────────────────────────────────────────
# 科目管理
# ────────────────────────────────────────────
@ms_bp.route("/subjects")
@require_role("ms_admin")
def subject_manage():
    subjects = Subject.query.order_by(Subject.sort_order).all()
    return render_template("ms/subject_manage.html", subjects=subjects)


@ms_bp.route("/subjects/create", methods=["GET", "POST"])
@require_role("ms_admin")
def subject_create():
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        full_score = request.form.get("full_score", 100, type=float)
        pass_score = request.form.get("pass_score", 60, type=float)
        sort_order = request.form.get("sort_order", 0, type=int)
        if not name:
            flash("科目名称不能为空", "warning")
            return redirect(url_for("ms.subject_manage"))
        if Subject.query.filter_by(name=name).first():
            flash("该科目已存在", "warning")
            return redirect(url_for("ms.subject_manage"))
        s = Subject(name=name, full_score=full_score,
                     pass_score=pass_score, sort_order=sort_order)
        db.session.add(s)
        safe_commit()
        flash("科目已创建", "success")
        return redirect(url_for("ms.subject_manage"))
    return render_template("ms/subject_form.html", subject=None)


@ms_bp.route("/subjects/<int:sid>/edit", methods=["GET", "POST"])
@require_role("ms_admin")
def subject_edit(sid):
    s = Subject.query.get_or_404(sid)
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        full_score = request.form.get("full_score", 100, type=float)
        pass_score = request.form.get("pass_score", 60, type=float)
        sort_order = request.form.get("sort_order", 0, type=int)
        if not name:
            flash("科目名称不能为空", "warning")
            return redirect(url_for("ms.subject_edit", sid=sid))
        exist = Subject.query.filter(Subject.name == name, Subject.id != sid).first()
        if exist:
            flash("该科目名称已被使用", "warning")
            return redirect(url_for("ms.subject_edit", sid=sid))
        s.name = name
        s.full_score = full_score
        s.pass_score = pass_score
        s.sort_order = sort_order
        safe_commit()
        flash("科目已更新", "success")
        return redirect(url_for("ms.subject_manage"))
    return render_template("ms/subject_form.html", subject=s)


@ms_bp.route("/subjects/<int:sid>/delete", methods=["POST"])
@require_role("ms_admin")
def subject_delete(sid):
    s = Subject.query.get_or_404(sid)
    s.is_active = False
    safe_commit()
    flash("科目已标记删除", "success")
    return redirect(url_for("ms.subject_manage"))


# ── 考勤总览（德育处全局视图） ──
@ms_bp.route("/attendance")
@require_role("ms_admin")
def attendance_overview():
    """德育处全局考勤历史：按年级/班级/日期范围筛选，支持导出"""
    grade_id = request.args.get("grade_id", type=int)
    class_id = request.args.get("class_id", type=int)
    start_date = request.args.get("start_date", "")
    end_date = request.args.get("end_date", "")
    export = request.args.get("export", "")

    today = date.today()
    grades = Grade.query.filter_by(is_active=True).order_by(Grade.sort_order).all()
    classes = Class.query.filter_by(is_active=True).order_by(Class.grade_id, Class.name).all()

    # 默认今天
    if not start_date:
        start_date = today.strftime("%Y-%m-%d")
    if not end_date:
        end_date = today.strftime("%Y-%m-%d")

    q = Attendance.query
    if grade_id:
        q = q.filter(Attendance.grade_id == grade_id)
    if class_id:
        q = q.filter(Attendance.class_id == class_id)

    try:
        sd = date.fromisoformat(start_date)
        ed = date.fromisoformat(end_date)
    except ValueError:
        sd = today
        ed = today
    q = q.filter(Attendance.record_date >= sd, Attendance.record_date <= ed)

    # 导出
    if export == "excel":
        import io
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "考勤记录"
        ws.append(["日期", "年级", "班级", "学号", "姓名", "状态", "备注"])
        records = q.order_by(Attendance.record_date.desc(), Attendance.grade_id,
                              Attendance.class_id, Attendance.student_id).all()
        for r in records:
            stu = Student.query.get(r.student_id) if r.student_id else None
            cls = Class.query.get(r.class_id) if r.class_id else None
            g = Grade.query.get(r.grade_id) if r.grade_id else None
            status_map = {"present": "出勤", "late": "迟到", "early": "早退",
                          "absent": "缺勤", "leave": "请假"}
            ws.append([
                str(r.record_date), g.name if g else "", cls.name if cls else "",
                stu.student_no if stu else "", stu.name if stu else "",
                status_map.get(r.status, r.status), r.note or ""
            ])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        from flask import send_file
        return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                         as_attachment=True, download_name=f"attendance_{start_date}_{end_date}.xlsx")

    # 按班级汇总统计
    summary_q = db.session.query(
        Attendance.class_id, Class.name,
        func.count().label("total"),
        func.sum(db.case((Attendance.status == "present", 1), else_=0)).label("present_cnt"),
        func.sum(db.case((Attendance.status == "late", 1), else_=0)).label("late_cnt"),
        func.sum(db.case((Attendance.status == "absent", 1), else_=0)).label("absent_cnt"),
        func.sum(db.case((Attendance.status == "leave", 1), else_=0)).label("leave_cnt"),
    ).join(Class, Attendance.class_id == Class.id)

    summary_q = summary_q.filter(Attendance.record_date >= sd, Attendance.record_date <= ed)
    if grade_id:
        summary_q = summary_q.filter(Attendance.grade_id == grade_id)
    if class_id:
        summary_q = summary_q.filter(Attendance.class_id == class_id)
    summary_q = summary_q.group_by(Attendance.class_id, Class.name).order_by(Class.grade_id, Class.name)
    summary = summary_q.all()

    # 分页明细
    page = request.args.get("page", 1, type=int)
    per_page = 50
    records = q.order_by(Attendance.record_date.desc(), Attendance.class_id).paginate(
        page=page, per_page=per_page, error_out=False)

    # 为每条记录附加学生姓名
    student_ids = list(set(r.student_id for r in records.items if r.student_id))
    students_map = {}
    if student_ids:
        for s in Student.query.filter(Student.id.in_(student_ids)).all():
            students_map[s.id] = s

    return render_template("ms/attendance.html",
                           grades=grades, classes=classes,
                           summary=summary, records=records,
                           students_map=students_map,
                           grade_filter=grade_id, class_filter=class_id,
                           start_date=start_date, end_date=end_date)


# ── 学生管理（德育处全局视图） ──
@ms_bp.route("/students")
@require_role("ms_admin")
def student_manage():
    """学生列表 - 支持按年级、班级、姓名/学号搜索"""
    grade_id = request.args.get("grade_id", type=int)
    class_id = request.args.get("class_id", type=int)
    q = request.args.get("q", "").strip()
    gender = request.args.get("gender", "")
    tag = request.args.get("tag", "").strip()
    page = request.args.get("page", 1, type=int)

    query = Student.query.filter_by(is_active=True)

    if grade_id:
        query = query.filter_by(grade_id=grade_id)
    if class_id:
        query = query.filter_by(class_id=class_id)
    if q:
        like = f"%{q}%"
        query = query.filter(
            (Student.name.like(like)) | (Student.student_no.like(like))
        )
    if gender:
        query = query.filter_by(gender=gender)
    if tag:
        like_tag = f"%{tag}%"
        query = query.filter(Student.tags.like(like_tag))

    query = query.order_by(Student.grade_id, Student.class_id, Student.student_no)
    pagination = query.paginate(page=page, per_page=50, error_out=False)

    grades = Grade.query.filter_by(is_active=True).order_by(Grade.sort_order).all()
    classes = Class.query.filter_by(is_active=True).order_by(Class.grade_id, Class.name).all()

    return render_template("ms/student_manage.html",
                           students=pagination.items, pagination=pagination,
                           grades=grades, classes=classes,
                           grade_filter=grade_id, class_filter=class_id,
                           q=q, gender=gender, tag=tag)


@ms_bp.route("/students/create", methods=["GET", "POST"])
@require_role("ms_admin")
@audit_log("add_student", "Student")
def student_create():
    """添加单个学生"""
    classes = Class.query.filter_by(is_active=True).order_by(Class.grade_id, Class.name).all()
    grades = Grade.query.filter_by(is_active=True).order_by(Grade.sort_order).all()

    if request.method == "POST":
        class_id = int(request.form["class_id"])
        cls = Class.query.get_or_404(class_id)
        grade_id = cls.grade_id

        s = Student(
            name=request.form["name"],
            student_no=request.form["student_no"],
            class_id=class_id,
            grade_id=grade_id,
            gender=request.form.get("gender", "男"),
            id_card=request.form.get("id_card", ""),
            national_id=request.form.get("national_id", ""),
            ethnicity=request.form.get("ethnicity", "汉族"),
            birth_date=datetime.strptime(request.form["birth_date"], "%Y-%m-%d").date() if request.form.get("birth_date") else None,
            address=request.form.get("address", ""),
            parent1_name=request.form.get("parent1_name", ""),
            parent1_phone=request.form.get("parent1_phone", ""),
            parent1_relation=request.form.get("parent1_relation", ""),
            parent2_name=request.form.get("parent2_name", ""),
            parent2_phone=request.form.get("parent2_phone", ""),
            parent2_relation=request.form.get("parent2_relation", ""),
            primary_school=request.form.get("primary_school", ""),
            tags=request.form.get("tags", ""),
        )
        db.session.add(s)
        safe_commit()
        flash("学生已添加", "success")
        return redirect(url_for("ms.student_manage"))

    return render_template("ms/student_form.html", student=None, action="create",
                           classes=classes, grades=grades)


@ms_bp.route("/students/<int:sid>/edit", methods=["GET", "POST"])
@require_role("ms_admin")
@audit_log("edit_student", "Student")
def student_edit(sid):
    """编辑学生"""
    student = Student.query.get_or_404(sid)
    classes = Class.query.filter_by(is_active=True).order_by(Class.grade_id, Class.name).all()
    grades = Grade.query.filter_by(is_active=True).order_by(Grade.sort_order).all()

    if request.method == "POST":
        class_id = int(request.form["class_id"])
        cls = Class.query.get_or_404(class_id)
        student.name = request.form["name"]
        student.student_no = request.form["student_no"]
        student.class_id = class_id
        student.grade_id = cls.grade_id
        student.gender = request.form.get("gender", "男")
        student.id_card = request.form.get("id_card", "")
        student.national_id = request.form.get("national_id", "")
        student.ethnicity = request.form.get("ethnicity", "汉族")
        student.birth_date = datetime.strptime(request.form["birth_date"], "%Y-%m-%d").date() if request.form.get("birth_date") else None
        student.address = request.form.get("address", "")
        student.parent1_name = request.form.get("parent1_name", "")
        student.parent1_phone = request.form.get("parent1_phone", "")
        student.parent1_relation = request.form.get("parent1_relation", "")
        student.parent2_name = request.form.get("parent2_name", "")
        student.parent2_phone = request.form.get("parent2_phone", "")
        student.parent2_relation = request.form.get("parent2_relation", "")
        student.primary_school = request.form.get("primary_school", "")
        student.tags = request.form.get("tags", "")
        safe_commit()
        flash("学生信息已更新", "success")
        return redirect(url_for("ms.student_manage"))

    return render_template("ms/student_form.html", student=student, action="edit",
                           classes=classes, grades=grades)


@ms_bp.route("/students/<int:sid>/delete", methods=["POST"])
@require_role("ms_admin")
@audit_log("delete_student", "Student")
def student_delete(sid):
    """软删除学生"""
    student = Student.query.get_or_404(sid)
    student.is_active = False
    safe_commit()
    flash("学生已删除", "success")
    return redirect(url_for("ms.student_manage"))


@ms_bp.route("/students/import", methods=["POST"])
@require_role("ms_admin")
def student_import():
    """Excel批量导入学生"""
    from io import BytesIO
    from openpyxl import load_workbook

    f = request.files.get("file")
    if not f or not f.filename:
        flash("请选择文件", "danger")
        return redirect(url_for("ms.student_manage"))

    class_id = request.form.get("class_id", type=int)
    if not class_id:
        flash("请选择班级", "danger")
        return redirect(url_for("ms.student_manage"))

    cls = Class.query.get_or_404(class_id)
    grade_id = cls.grade_id

    wb = load_workbook(BytesIO(f.read()))
    ws = wb.active

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        no = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] else ""
        if not name:
            continue
        if Student.query.filter_by(student_no=no).first():
            continue  # 跳过重复学号
        s = Student(
            student_no=no,
            name=name,
            class_id=class_id,
            grade_id=grade_id,
            gender=str(row[2]).strip() if row[2] else "男",
            id_card=str(row[3]).strip() if row[3] else "",
            national_id=str(row[4]).strip() if row[4] else "",
            parent1_name=str(row[5]).strip() if row[5] else "",
            parent1_phone=str(row[6]).strip() if row[6] else "",
        )
        db.session.add(s)
        count += 1

    safe_commit()
    flash(f"成功导入 {count} 名学生", "success")
    return redirect(url_for("ms.student_manage"))


@ms_bp.route("/students/template")
@require_role("ms_admin")
def student_template():
    """下载Excel导入模板"""
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "学生导入模板"
    ws.append(["学号*", "姓名*", "性别", "身份证号", "全国学籍号", "家长1姓名", "家长1电话"])
    ws.append(["2024001", "张三", "男", "430XXXXXXX", "GXXXXXXXX", "张三父", "13800000000"])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from flask import send_file
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="student_template.xlsx")


@ms_bp.route("/students/export")
@require_role("ms_admin")
def student_export():
    """导出学生到Excel"""
    from openpyxl import Workbook
    from io import BytesIO

    grade_id = request.args.get("grade_id", type=int)
    class_id = request.args.get("class_id", type=int)

    query = Student.query.filter_by(is_active=True)
    if grade_id:
        query = query.filter_by(grade_id=grade_id)
    if class_id:
        query = query.filter_by(class_id=class_id)

    students = query.order_by(Student.grade_id, Student.class_id, Student.student_no).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "学生名册"
    ws.append(["年级", "班级", "学号", "姓名", "性别", "家长1", "电话1", "家长2", "电话2", "标签"])

    for s in students:
        cls = Class.query.get(s.class_id)
        grade = Grade.query.get(s.grade_id)
        ws.append([
            grade.name if grade else "",
            cls.name if cls else "",
            s.student_no, s.name, s.gender,
            s.parent1_name, s.parent1_phone,
            s.parent2_name, s.parent2_phone,
            s.tags
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from flask import send_file
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="students_export.xlsx")

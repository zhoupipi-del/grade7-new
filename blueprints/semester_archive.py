"""学期归档 — 学期数据快照/查看/恢复/对比"""
from datetime import datetime
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from models import db, SemesterArchive, Student, DisciplineRecord, Score, Activity, Message, Attendance
from decorators import login_required, require_role
from utils.db_utils import safe_commit

semester_archive_bp = Blueprint("semester_archive", __name__, template_folder="../templates")


@semester_archive_bp.route("/")
@login_required
def archive_list():
    """归档列表"""
    archives = SemesterArchive.query.order_by(SemesterArchive.archived_at.desc()).all()
    # 生成当前学期名称
    now = datetime.now()
    if now.month >= 9:
        current_semester = f"{now.year}-{now.year+1}-1"
    elif now.month >= 3:
        current_semester = f"{now.year-1}-{now.year}-2"
    else:
        current_semester = f"{now.year-1}-{now.year}-1"
    return render_template("semester_archive/list.html", archives=archives, current_semester=current_semester)


@semester_archive_bp.route("/<semester_name>")
@login_required
def archive_detail(semester_name):
    """查看归档详情"""
    archive = SemesterArchive.query.filter_by(semester_name=semester_name).first_or_404()
    summary = archive.summary_json if archive.summary_json else "{}"
    import json
    data = json.loads(summary)
    return render_template("semester_archive/detail.html", archive=archive, data=data)


@semester_archive_bp.route("/<semester_name>/create", methods=["POST"])
@login_required
@require_role("ms_admin")
def create_archive(semester_name):
    """创建学期归档"""
    if SemesterArchive.query.filter_by(semester_name=semester_name).first():
        flash(f"学期 [{semester_name}] 的归档已存在", "warning")
        return redirect(url_for("semester_archive.archive_list"))

    summary = _collect_semester_data(semester_name)
    import json
    archive = SemesterArchive(
        semester_name=semester_name,
        display_name=request.form.get("display_name", semester_name),
        start_date=datetime.strptime(request.form.get("start_date", "2025-09-01"), "%Y-%m-%d").date(),
        end_date=datetime.strptime(request.form.get("end_date", "2026-01-31"), "%Y-%m-%d").date(),
        summary_json=json.dumps(summary, ensure_ascii=False, default=str),
    )
    db.session.add(archive)
    safe_commit()
    flash(f"学期 [{semester_name}] 归档创建成功", "success")
    return redirect(url_for("semester_archive.archive_list"))


@semester_archive_bp.route("/<semester_name>/delete", methods=["POST"])
@login_required
@require_role("ms_admin")
def delete_archive(semester_name):
    """删除归档"""
    archive = SemesterArchive.query.filter_by(semester_name=semester_name).first_or_404()
    db.session.delete(archive)
    safe_commit()
    flash(f"归档 [{semester_name}] 已删除", "success")
    return redirect(url_for("semester_archive.archive_list"))


@semester_archive_bp.route("/<semester_name>/restore", methods=["POST"])
@login_required
@require_role("ms_admin")
def restore_archive(semester_name):
    """从归档恢复数据（覆盖当前学期部分数据）"""
    archive = SemesterArchive.query.filter_by(semester_name=semester_name).first_or_404()
    import json
    data = json.loads(archive.summary_json) if archive.summary_json else {}
    # 恢复数据逻辑（TODO：根据 summary_json 恢复对应表数据）
    flash(f"已从归档 [{semester_name}] 恢复数据（功能完善中）", "warning")
    return redirect(url_for("semester_archive.archive_list"))


@semester_archive_bp.route("/compare")
@login_required
def compare_semesters():
    """学期对比页面"""
    archives = SemesterArchive.query.order_by(SemesterArchive.semester_name).all()
    return render_template("semester_archive/compare.html", archives=archives)


@semester_archive_bp.route("/api/compare/<s1>/<s2>")
@login_required
def api_compare(s1, s2):
    """API：对比两个学期数据"""
    import json
    archive1 = SemesterArchive.query.filter_by(semester_name=s1).first()
    archive2 = SemesterArchive.query.filter_by(semester_name=s2).first()

    if not archive1 or not archive2:
        return jsonify({"error": "归档不存在"}), 404

    data1 = json.loads(archive1.summary_json) if archive1.summary_json else {}
    data2 = json.loads(archive2.summary_json) if archive2.summary_json else {}

    def _calc_change(a, b):
        if a == 0:
            return "+100%" if b > 0 else "0%"
        return f"{'+' if b >= a else ''}{((b - a) / a * 100):.1f}%"

    comparison = {
        "semester1": s1,
        "semester2": s2,
        "changes": {
            "student_count": _calc_change(data1.get("student_count", 0), data2.get("student_count", 0)),
            "discipline_total": _calc_change(data1.get("discipline_stats", {}).get("total", 0), data2.get("discipline_stats", {}).get("total", 0)),
            "average_score": _calc_change(data1.get("score_stats", {}).get("average", 0), data2.get("score_stats", {}).get("average", 0)),
            "activity_participation": _calc_change(data1.get("activity_stats", {}).get("total_participants", 0), data2.get("activity_stats", {}).get("total_participants", 0)),
        }
    }
    return jsonify(comparison)


@semester_archive_bp.route("/<semester_name>/export")
@login_required
def export_archive(semester_name):
    """导出归档数据为Excel"""
    import json
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    archive = SemesterArchive.query.filter_by(semester_name=semester_name).first_or_404()
    data = json.loads(archive.summary_json) if archive.summary_json else {}

    wb = Workbook()
    ws = wb.active
    ws.title = "归档数据"

    header_font = Font(size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    def write_section(ws, row, title, data_dict, col_count=2):
        ws.cell(row=row, column=1, value=title).font = Font(size=13, bold=True)
        row += 1
        for key, value in data_dict.items():
            if isinstance(value, dict):
                ws.cell(row=row, column=1, value=key).font = Font(bold=True)
                row += 1
                for sub_key, sub_value in value.items():
                    ws.cell(row=row, column=1, value=sub_key)
                    ws.cell(row=row, column=2, value=str(sub_value) if sub_value is not None else "")
                    row += 1
            else:
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=str(value) if value is not None else "")
                row += 1
        return row + 1

    row = 1
    ws.cell(row=row, column=1, value=f"学期归档: {archive.display_name}").font = Font(size=16, bold=True)
    row += 2

    for section in ["student_count", "class_count", "discipline_stats", "score_stats", "activity_stats", "message_stats"]:
        if section in data:
            row = write_section(ws, row, section, {section: data[section]} if not isinstance(data[section], dict) else data[section])

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from flask import send_file
    return send_file(
        output,
        as_attachment=True,
        download_name=f"归档_{semester_name}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def _collect_semester_data(semester_name):
    """收集学期数据（快照）"""
    from sqlalchemy import func
    summary = {}

    # 学生总数
    summary["student_count"] = Student.query.count()

    # 班级数
    from models import Class
    summary["class_count"] = Class.query.count()

    # 违纪统计
    disc_stats = {}
    disc_total = DisciplineRecord.query.count()
    disc_active = DisciplineRecord.query.filter_by(status="active").count()
    disc_resolved = DisciplineRecord.query.filter_by(status="resolved").count()
    disc_stats["total"] = disc_total
    disc_stats["active"] = disc_active
    disc_stats["resolved"] = disc_resolved
    summary["discipline_stats"] = disc_stats

    # 成绩统计
    score_stats = {}
    avg_score = db.session.query(func.avg(Score.score)).scalar() or 0
    score_stats["average"] = float(avg_score)
    summary["score_stats"] = score_stats

    # 活动统计
    activity_stats = {}
    activity_stats["total"] = Activity.query.count()
    activity_stats["total_participants"] = db.session.query(func.count(Activity.participants)).scalar() or 0
    summary["activity_stats"] = activity_stats

    # 消息统计
    message_stats = {}
    message_stats["total"] = Message.query.count()
    message_stats["unread"] = Message.query.filter_by(is_read=False).count()
    summary["message_stats"] = message_stats

    return summary


def _calc_change(a, b):
    """计算变化百分比"""
    if a == 0:
        return "+100%" if b > 0 else "0%"
    diff = ((b - a) / a) * 100
    sign = "+" if diff >= 0 else ""
    return f"{sign}{diff:.1f}%"

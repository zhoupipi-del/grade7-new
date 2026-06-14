"""导出汇总 — 各模块Excel导出"""
from flask import Blueprint, render_template, request, send_file, session
from datetime import datetime
from io import BytesIO
import openpyxl

from decorators import login_required, require_role
from models import db, Student, DisciplineRecord, RoutineScore, Task, WingsScore, HomeVisit, EndTermComment, Class, Grade, User, Score, Exam, Subject, Attendance

export_summary_bp = Blueprint("export_summary", __name__, url_prefix="/export-summary")


EXPORT_MODULES = [
    {"key": "students", "name": "学生花名册", "icon": "bi-people"},
    {"key": "discipline", "name": "违纪记录", "icon": "bi-exclamation-triangle"},
    {"key": "routine", "name": "常规评分", "icon": "bi-star"},
    {"key": "tasks", "name": "德育任务", "icon": "bi-list-check"},
    {"key": "wings", "name": "五翼评价", "icon": "bi-graph-up"},
    {"key": "scores", "name": "成绩汇总", "icon": "bi-journal-check"},
    {"key": "attendance", "name": "考勤汇总", "icon": "bi-calendar-check"},
    {"key": "home_visits", "name": "家访记录", "icon": "bi-house"},
    {"key": "comments", "name": "期末评语", "icon": "bi-chat-text"},
]


@export_summary_bp.route("/")
@login_required
@require_role("ms_admin", "grade_leader")
def index():
    """导出汇总页"""
    user = db.session.get(User, session.get("user_id"))
    grade_id = user.grade_id if user.role == "grade_leader" else None

    counts = {}
    if grade_id:
        counts["students"] = Student.query.filter_by(grade_id=grade_id, is_active=True).count()
        counts["discipline"] = DisciplineRecord.query.filter_by(grade_id=grade_id).count()
        counts["routine"] = RoutineScore.query.filter_by(grade_id=grade_id).count()
        counts["tasks"] = Task.query.count()
        counts["wings"] = WingsScore.query.join(Student).filter(Student.grade_id == grade_id).count()
        counts["scores"] = Score.query.filter_by(grade_id=grade_id).count()
        counts["attendance"] = Attendance.query.filter_by(grade_id=grade_id).count()
        counts["home_visits"] = HomeVisit.query.filter_by(grade_id=grade_id).count()
        counts["comments"] = EndTermComment.query.filter_by(grade_id=grade_id).count()
    else:
        counts["students"] = Student.query.filter_by(is_active=True).count()
        counts["discipline"] = DisciplineRecord.query.count()
        counts["routine"] = RoutineScore.query.count()
        counts["tasks"] = Task.query.count()
        counts["wings"] = WingsScore.query.count()
        counts["scores"] = Score.query.count()
        counts["attendance"] = Attendance.query.count()
        counts["home_visits"] = HomeVisit.query.count()
        counts["comments"] = EndTermComment.query.count()

    grades = Grade.query.all() if not grade_id else []

    return render_template("export_summary/index.html",
                           modules=EXPORT_MODULES, counts=counts,
                           grades=grades, grade_id=grade_id)


@export_summary_bp.route("/excel/<module_key>")
@login_required
@require_role("ms_admin", "grade_leader")
def excel(module_key):
    """导出指定模块Excel（支持逗号分隔多个模块，生成多Sheet文件）"""
    user = db.session.get(User, session.get("user_id"))
    grade_id = user.grade_id if user.role == "grade_leader" else None

    wb = openpyxl.Workbook()
    # 删除默认创建的空Sheet（将在各导出函数中创建命名Sheet）
    wb.remove(wb.active)

    # 支持逗号分隔的多个模块
    module_keys = [k.strip() for k in module_key.split(",") if k.strip()]

    for idx, key in enumerate(module_keys):
        if key == "students":
            _export_students(wb, grade_id)
        elif key == "discipline":
            _export_discipline(wb, grade_id)
        elif key == "routine":
            _export_routine(wb, grade_id)
        elif key == "tasks":
            _export_tasks(wb, grade_id)
        elif key == "wings":
            _export_wings(wb, grade_id)
        elif key == "scores":
            _export_scores(wb, grade_id)
        elif key == "attendance":
            _export_attendance(wb, grade_id)
        elif key == "home_visits":
            _export_visits(wb, grade_id)
        elif key == "comments":
            _export_comments(wb, grade_id)

    # 如果所有模块都无效，wb 没有 Sheet，需要添加一个错误提示
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet("错误")
        ws.append(["未知模块", module_key])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    name = "_".join([m["name"] for m in EXPORT_MODULES if m["key"] in module_keys]) or "导出"
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"{name}_{datetime.now().strftime('%Y%m%d')}.xlsx")


def _safe_creator(rec):
    """安全获取创建者名称"""
    if rec.creator:
        return rec.creator.display_name
    return ""


def _export_students(wb, grade_id=None):
    """导出学生花名册到独立Sheet"""
    ws = wb.create_sheet(title="学生花名册")
    ws.append(["学号", "姓名", "性别", "班级", "民族", "家长姓名", "家长电话", "标签"])
    q = Student.query.filter_by(is_active=True).order_by(Student.class_id, Student.student_no)
    if grade_id:
        q = q.filter_by(grade_id=grade_id)
    for s in q.all():
        ws.append([s.student_no, s.name, s.gender,
                   s.class_.name if s.class_ else "",
                   s.ethnicity, s.parent1_name or "", s.parent1_phone or "", s.tags])


def _export_discipline(wb, grade_id=None):
    """导出违纪记录到独立Sheet"""
    ws = wb.create_sheet(title="违纪记录")
    ws.append(["日期", "学生姓名", "班级", "违纪类别", "等级", "描述", "登记人"])
    q = DisciplineRecord.query.order_by(DisciplineRecord.created_at.desc())
    if grade_id:
        q = q.filter_by(grade_id=grade_id)
    for d in q.all():
        ws.append([d.created_at.strftime("%Y-%m-%d %H:%M") if d.created_at else "",
                   d.student.name if d.student else "",
                   d.student.class_.name if d.student and d.student.class_ else "",
                   d.category or "", d.type, d.description, _safe_creator(d)])


def _export_routine(wb, grade_id=None):
    """导出常规评分到独立Sheet"""
    ws = wb.create_sheet(title="常规评分")
    ws.append(["日期", "班级", "类别", "分数", "备注", "检查人"])
    q = RoutineScore.query.order_by(RoutineScore.record_date.desc())
    if grade_id:
        q = q.filter_by(grade_id=grade_id)
    for r in q.all():
        ws.append([r.record_date.strftime("%Y-%m-%d") if r.record_date else "",
                   r.class_.name if r.class_ else "",
                   r.category, r.score, r.note or "", r.inspector or ""])


def _export_tasks(wb, grade_id=None):
    """导出德育任务到独立Sheet"""
    ws = wb.create_sheet(title="德育任务")
    ws.append(["标题", "内容", "来源角色", "状态", "创建时间", "截止时间"])
    q = Task.query.order_by(Task.created_at.desc())
    # tasks don't have grade_id; use from_user's grade if needed
    for t in q.all():
        ws.append([t.title, t.content or "", t.from_role,
                   t.status,
                   t.created_at.strftime("%Y-%m-%d %H:%M") if t.created_at else "",
                   t.deadline.strftime("%Y-%m-%d") if t.deadline else ""])


def _export_wings(wb, grade_id=None):
    """导出五翼评价到独立Sheet"""
    ws = wb.create_sheet(title="五翼评价")
    ws.append(["学生姓名", "班级", "评分维度", "分数", "来源", "评分人", "日期"])
    q = WingsScore.query.order_by(WingsScore.created_at.desc())
    if grade_id:
        q = q.join(Student).filter(Student.grade_id == grade_id)
    for w in q.all():
        ws.append([w.student.name if w.student else "",
                   w.student.class_.name if w.student and w.student.class_ else "",
                   w.dimension, w.score, w.scorer_type, str(w.scorer_id),
                   w.created_at.strftime("%Y-%m-%d %H:%M") if w.created_at else ""])


def _export_visits(wb, grade_id=None):
    """导出家访记录到独立Sheet"""
    ws = wb.create_sheet(title="家访记录")
    ws.append(["日期", "学生姓名", "班级", "家访方式", "教师", "内容", "家长反馈"])
    q = HomeVisit.query.order_by(HomeVisit.visit_date.desc())
    if grade_id:
        q = q.filter_by(grade_id=grade_id)
    for v in q.all():
        ws.append([v.visit_date.strftime("%Y-%m-%d"),
                   v.student.name if v.student else "",
                   v.student.class_.name if v.student and v.student.class_ else "",
                   v.visit_type, v.teacher_name, v.content_summary, v.parent_feedback])


def _export_comments(wb, grade_id=None):
    """导出期末评语到独立Sheet"""
    ws = wb.create_sheet(title="期末评语")
    ws.append(["学生姓名", "班级", "学期", "综合评语", "优点", "待改进", "状态"])
    q = EndTermComment.query.order_by(EndTermComment.updated_at.desc())
    if grade_id:
        q = q.filter_by(grade_id=grade_id)
    for c in q.all():
        ws.append([c.student.name if c.student else "",
                   c.student.class_.name if c.student and c.student.class_ else "",
                   c.semester, c.overall_comment, c.strengths, c.improvements, c.status])


def _export_scores(wb, grade_id=None):
    """导出成绩汇总到独立Sheet"""
    ws = wb.create_sheet(title="成绩汇总")
    ws.append(["考试名称", "考试日期", "学生姓名", "班级", "科目", "分数", "班级排名", "年级排名"])
    q = Score.query.join(Exam).order_by(Exam.exam_date.desc(), Score.student_id, Score.subject_id)
    if grade_id:
        q = q.filter(Score.grade_id == grade_id)
    for s in q.all():
        ws.append([s.exam.name if s.exam else "",
                   s.exam.exam_date.strftime("%Y-%m-%d") if s.exam and s.exam.exam_date else "",
                   s.student.name if s.student else "",
                   s.student.class_.name if s.student and s.student.class_ else "",
                   s.subject.name if s.subject else "",
                   s.score, s.rank_class, s.rank_grade])


def _export_attendance(wb, grade_id=None):
    """导出考勤汇总到独立Sheet"""
    ws = wb.create_sheet(title="考勤汇总")
    ws.append(["日期", "学生姓名", "班级", "状态", "备注"])
    q = Attendance.query.order_by(Attendance.record_date.desc())
    if grade_id:
        q = q.filter_by(grade_id=grade_id)
    status_labels = {"present": "出勤", "late": "迟到", "early": "早退", "absent": "缺勤", "leave": "请假"}
    for a in q.all():
        ws.append([a.record_date.strftime("%Y-%m-%d") if a.record_date else "",
                   a.student.name if a.student else "",
                   a.student.class_.name if a.student and a.student.class_ else "",
                   status_labels.get(a.status, a.status), a.note or ""])

"""报表自动生成 — 使用 openpyxl 生成 Excel 报表"""
import os
import json
from datetime import datetime, date, timedelta
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, jsonify, session
from models import db, Report, Student, DisciplineRecord, Score, Activity, Message, Attendance, Grade, Class
from decorators import login_required, require_role
from utils.db_utils import safe_commit

report_generator_bp = Blueprint("report_generator", __name__, template_folder="../templates")

REPORT_DIR = "static/reports"


def _ensure_report_dir():
    """确保报表目录存在"""
    path = os.path.join(os.path.dirname(__file__), "..", REPORT_DIR)
    os.makedirs(path, exist_ok=True)
    return path


@report_generator_bp.route("/")
@login_required
def report_list():
    """报表列表"""
    page = request.args.get("page", 1, type=int)
    pagination = Report.query.order_by(Report.generated_at.desc()).paginate(page=page, per_page=20, error_out=False)
    return render_template("report_generator/list.html", pagination=pagination)


@report_generator_bp.route("/generate", methods=["GET", "POST"])
@login_required
@require_role("ms_admin", "grade_leader")
def generate_report():
    """生成报表"""
    if request.method == "POST":
        report_type = request.form.get("report_type", "")
        title = request.form.get("title", "")
        semester = request.form.get("semester", "")
        grade_id = request.form.get("grade_id", type=int)
        class_id = request.form.get("class_id", type=int)

        if not report_type or not title:
            flash("请填写报表类型和标题", "warning")
            return redirect(url_for("report_generator.generate_report"))

        # 创建报表记录
        report = Report(
            report_type=report_type,
            title=title,
            semester=semester,
            grade_id=grade_id if grade_id else None,
            class_id=class_id if class_id else None,
            generated_by_id=session.get("user_id"),
        )
        db.session.add(report)
        safe_commit()

        # 根据类型生成报表数据
        data = {}
        if report_type == "weekly":
            data = _generate_weekly_data(semester, grade_id, class_id)
        elif report_type == "monthly":
            data = _generate_monthly_data(semester, grade_id, class_id)
        elif report_type == "semester":
            data = _generate_semester_data(semester, grade_id, class_id)
        elif report_type == "custom":
            start_date = request.form.get("start_date")
            end_date = request.form.get("end_date")
            data = _generate_custom_data(start_date, end_date, grade_id, class_id)

        report.data_json = json.dumps(data, ensure_ascii=False, default=str)
        safe_commit()

        flash(f"报表 [{title}] 生成成功", "success")
        return redirect(url_for("report_generator.view_report", rid=report.id))

    return render_template("report_generator/generate.html")


@report_generator_bp.route("/<int:rid>/view")
@login_required
def view_report(rid):
    """查看报表详情"""
    report = Report.query.get_or_404(rid)
    data = json.loads(report.data_json) if report.data_json else {}
    return render_template("report_generator/view.html", report=report, data=data)


@report_generator_bp.route("/<int:rid>/download")
@login_required
def download_report(rid):
    """下载报表 Excel 文件"""
    report = Report.query.get_or_404(rid)

    if report.file_path and os.path.exists(report.file_path):
        return send_file(report.file_path, as_attachment=True)

    # 如果没有生成 Excel，先生成
    data = json.loads(report.data_json) if report.data_json else {}
    file_path = _generate_excel(report, data)

    if file_path:
        return send_file(file_path, as_attachment=True)
    else:
        flash("报表文件生成失败", "danger")
        return redirect(url_for("report_generator.view_report", rid=rid))


@report_generator_bp.route("/<int:rid>/delete", methods=["POST"])
@login_required
@require_role("ms_admin")
def delete_report(rid):
    """删除报表"""
    report = Report.query.get_or_404(rid)

    # 删除文件
    if report.file_path and os.path.exists(report.file_path):
        os.remove(report.file_path)

    db.session.delete(report)
    safe_commit()
    flash(f"报表 [{report.title}] 已删除", "success")
    return redirect(url_for("report_generator.report_list"))


# ── 数据生成函数 ────────────────────────────────────────────────────────────────

def _generate_weekly_data(semester, grade_id=None, class_id=None):
    """生成周报表数据"""
    from datetime import timedelta
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)

    data = {
        "period": f"{week_start} ~ {week_end}",
        "discipline_stats": _get_discipline_stats(week_start, week_end, grade_id, class_id),
        "score_stats": _get_score_stats(semester, grade_id, class_id),
        "message_stats": _get_message_stats(week_start, week_end),
        "activity_stats": _get_activity_stats(week_start, week_end, grade_id, class_id),
    }
    return data


def _generate_monthly_data(semester, grade_id=None, class_id=None):
    """生成月报表数据"""
    today = date.today()
    month_start = date(today.year, today.month, 1)

    data = {
        "period": f"{today.year}年{today.month}月",
        "discipline_ranking": _get_discipline_ranking(grade_id, class_id),
        "score_ranking": _get_score_ranking(semester, grade_id),
        "top_discipline_students": _get_top_discipline_students(grade_id, class_id),
        "teacher_feedback_stats": _get_teacher_feedback_stats(month_start, grade_id),
    }
    return data


def _generate_semester_data(semester, grade_id=None, class_id=None):
    """生成学期报表数据"""
    data = {
        "semester": semester,
        "quality_evaluation": _get_quality_evaluation(semester, grade_id, class_id),
        "score_analysis": _get_score_analysis(semester, grade_id),
        "activity_participation": _get_activity_participation(semester, grade_id, class_id),
        "parent_interaction": _get_parent_interaction(semester, grade_id, class_id),
    }
    return data


def _generate_custom_data(start_date, end_date, grade_id=None, class_id=None):
    """生成自定义报表数据"""
    data = {
        "period": f"{start_date} ~ {end_date}",
        "discipline_stats": _get_discipline_stats(start_date, end_date, grade_id, class_id),
        "score_stats": {},
        "message_stats": _get_message_stats(start_date, end_date),
        "activity_stats": _get_activity_stats(start_date, end_date, grade_id, class_id),
    }
    return data


# ── 统计辅助函数 ────────────────────────────────────────────────────────────────

def _get_discipline_stats(start_date, end_date, grade_id=None, class_id=None):
    """获取违纪统计"""
    query = DisciplineRecord.query.filter(DisciplineRecord.date.between(start_date, end_date))
    if class_id:
        query = query.filter_by(class_id=class_id)
    elif grade_id:
        # 需要join Student来获取grade_id
        pass

    total = query.count()
    by_type = {}
    for d in query.all():
        by_type[d.type] = by_type.get(d.type, 0) + 1

    return {"total": total, "by_type": by_type}


def _get_score_stats(semester=None, grade_id=None, class_id=None):
    """获取成绩统计"""
    query = Score.query
    if semester:
        query = query.filter_by(semester=semester)
    if class_id:
        query = query.filter_by(class_id=class_id)
    elif grade_id:
        students = Student.query.filter_by(grade_id=grade_id).with_entities(Student.id).all()
        student_ids = [s[0] for s in students]
        if student_ids:
            query = query.filter(Score.student_id.in_(student_ids))

    scores = query.all()
    total = len(scores)
    avg = round(sum(s.score for s in scores if s.score) / total, 2) if total > 0 else 0

    by_subject = {}
    for s in scores:
        if s.score:
            subj = s.subject or "未知"
            if subj not in by_subject:
                by_subject[subj] = {"count": 0, "total": 0}
            by_subject[subj]["count"] += 1
            by_subject[subj]["total"] += s.score

    for subj, data in by_subject.items():
        data["average"] = round(data["total"] / data["count"], 2) if data["count"] > 0 else 0

    return {"total": total, "average": avg, "by_subject": by_subject}


def _get_message_stats(start_date, end_date):
    """获取消息统计"""
    query = Message.query
    if start_date and end_date:
        query = query.filter(Message.created_at.between(start_date, end_date))
    total = query.count()
    unread_total = Message.query.filter_by(is_read=False).count()
    return {"total": total, "unread": unread_total}


def _get_activity_stats(start_date, end_date, grade_id=None, class_id=None):
    """获取活动统计"""
    query = Activity.query
    if start_date and end_date:
        query = query.filter(Activity.start_time.between(start_date, end_date))
    if class_id:
        query = query.filter_by(class_id=class_id)
    elif grade_id:
        query = query.filter_by(grade_id=grade_id)

    activities = query.all()
    total = len(activities)
    total_participants = sum(a.current_participants or 0 for a in activities)
    by_status = {}
    for a in activities:
        status = a.status or "未知"
        by_status[status] = by_status.get(status, 0) + 1
    return {"total": total, "participants": total_participants, "by_status": by_status}


def _get_discipline_ranking(grade_id=None, class_id=None):
    """获取违纪排名（按班级）"""
    query = db.session.query(
        Student.class_,
        db.func.count(DisciplineRecord.id).label("count")
    ).join(DisciplineRecord, Student.id == DisciplineRecord.student_id)

    if class_id:
        query = query.filter(Student.class_id == class_id)
    elif grade_id:
        query = query.filter(Student.grade_id == grade_id)

    results = query.group_by(Student.class_).order_by(db.text("count DESC")).all()
    return [{"class": r[0] if r[0] else "未知", "count": int(r[1])} for r in results]


def _get_score_ranking(semester, grade_id=None):
    """获取成绩排名（按班级）"""
    query = db.session.query(
        Student.class_,
        db.func.avg(Score.score).label("avg_score")
    ).join(Score, Student.id == Score.student_id)

    if semester:
        query = query.filter(Score.semester == semester)
    if grade_id:
        query = query.filter(Student.grade_id == grade_id)

    results = query.group_by(Student.class_).order_by(db.text("avg_score DESC")).all()
    return [{"class": r[0] if r[0] else "未知", "average": round(float(r[1]), 2)} for r in results]


def _get_top_discipline_students(grade_id=None, class_id=None, limit=20):
    """获取违纪学生Top"""
    query = db.session.query(
        Student.name,
        Student.class_,
        db.func.count(DisciplineRecord.id).label("count")
    ).join(DisciplineRecord, Student.id == DisciplineRecord.student_id)

    if class_id:
        query = query.filter(Student.class_id == class_id)
    elif grade_id:
        query = query.filter(Student.grade_id == grade_id)

    results = query.group_by(Student.id).order_by(db.text("count DESC")).limit(limit).all()
    return [{"name": r[0], "class": r[1], "count": int(r[2])} for r in results]


def _get_teacher_feedback_stats(month_start, grade_id=None):
    """获取教师反馈统计（基于日常评分体系得分）"""
    return {"total_feedback": 0, "by_wing": {}}


def _get_quality_evaluation(semester, grade_id=None, class_id=None):
    """获取综合素质评价"""
    return {"total_evaluations": 0, "by_indicator": {}}


def _get_score_analysis(semester, grade_id=None):
    """获取成绩分析"""
    return {"pass_rate": 0, "excellent_rate": 0}


def _get_activity_participation(semester, grade_id=None, class_id=None):
    """获取活动参与统计"""
    query = Activity.query
    if semester and hasattr(Activity, 'semester'):
        query = query.filter(Activity.semester == semester)
    if class_id:
        query = query.filter_by(class_id=class_id)
    elif grade_id:
        query = query.filter_by(grade_id=grade_id)

    activities = query.all()
    result = {
        "total_activities": len(activities),
        "total_participants": sum(a.current_participants or 0 for a in activities),
        "by_type": {}
    }
    for a in activities:
        atype = a.type or "未知"
        if atype not in result["by_type"]:
            result["by_type"][atype] = {"count": 0, "participants": 0}
        result["by_type"][atype]["count"] += 1
        result["by_type"][atype]["participants"] += (a.current_participants or 0)
    return result


def _get_parent_interaction(semester, grade_id=None, class_id=None):
    """获取家长互动统计"""
    return {"total_messages": 0, "total_replies": 0}


def _generate_excel(report, data):
    """生成 Excel 文件（支持多sheet）"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.chart import PieChart, BarChart, Reference
        from openpyxl.utils import get_column_letter

        report_dir = _ensure_report_dir()
        filename = f"report_{report.id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        file_path = os.path.join(report_dir, filename)

        wb = Workbook()
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

        def style_header(ws, row, cols):
            for col in range(1, cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border

        # ── Sheet 1: 基本信息 ──
        ws1 = wb.active
        ws1.title = "概要"
        ws1['A1'] = report.title
        ws1['A1'].font = Font(size=18, bold=True)
        ws1.merge_cells('A1:D1')
        ws1['A3'] = "报表类型"; ws1['B3'] = report.report_type
        ws1['A4'] = "学期"; ws1['B4'] = report.semester or "默认"
        ws1['A5'] = "生成时间"; ws1['B5'] = report.generated_at.strftime("%Y-%m-%d %H:%M:%S")
        ws1['A6'] = "生成人"; ws1['B6'] = session.get("display_name", "")

        # ── Sheet 2: 违纪统计 ──
        if "discipline_stats" in data and data["discipline_stats"]:
            ds = data["discipline_stats"]
            ws2 = wb.create_sheet("违纪统计")
            ws2['A1'] = "违纪统计"
            ws2['A1'].font = Font(size=14, bold=True)
            ws2['A3'] = "违纪总数"; ws2['B3'] = ds.get("total", 0)

            ws2['A5'] = "违纪类型"; ws2['B5'] = "次数"
            style_header(ws2, 5, 2)
            row = 6
            by_type = ds.get("by_type", {})
            for typ, count in sorted(by_type.items(), key=lambda x: x[1], reverse=True):
                ws2.cell(row=row, column=1, value=typ)
                ws2.cell(row=row, column=2, value=count)
                row += 1

            # 饼图
            if by_type and len(by_type) <= 10:
                chart = PieChart()
                chart.title = "违纪类型分布"
                labels = Reference(ws2, min_col=1, min_row=6, max_row=row-1)
                values = Reference(ws2, min_col=2, min_row=6, max_row=row-1)
                chart.add_data(values, titles_from_data=False)
                chart.set_categories(labels)
                ws2.add_chart(chart, f"D3")

        # ── Sheet 3: 成绩统计 ──
        if "score_stats" in data and data["score_stats"]:
            ss = data["score_stats"]
            ws3 = wb.create_sheet("成绩统计")
            ws3['A1'] = "成绩统计"
            ws3['A1'].font = Font(size=14, bold=True)
            ws3['A3'] = "成绩总数"; ws3['B3'] = ss.get("total", 0)
            ws3['A4'] = "平均分"; ws3['B4'] = ss.get("average", 0)

            by_subject = ss.get("by_subject", {})
            if by_subject:
                ws3['A6'] = "科目"; ws3['B6'] = "人数"; ws3['C6'] = "平均分"
                style_header(ws3, 6, 3)
                row = 7
                for subj, info in by_subject.items():
                    ws3.cell(row=row, column=1, value=subj)
                    ws3.cell(row=row, column=2, value=info.get("count", 0))
                    ws3.cell(row=row, column=3, value=info.get("average", 0))
                    row += 1

        # ── Sheet 4: 违纪排名 ──
        if "discipline_ranking" in data and data["discipline_ranking"]:
            dr = data["discipline_ranking"]
            ws4 = wb.create_sheet("违纪排名")
            ws4['A1'] = "班级违纪排名"
            ws4['A1'].font = Font(size=14, bold=True)
            ws4['A3'] = "班级"; ws4['B3'] = "违纪次数"
            style_header(ws4, 3, 2)
            row = 4
            for item in dr:
                ws4.cell(row=row, column=1, value=item.get("class", ""))
                ws4.cell(row=row, column=2, value=item.get("count", 0))
                row += 1

            # 柱状图
            if len(dr) <= 15:
                chart = BarChart()
                chart.title = "班级违纪对比"
                chart.y_axis.title = "违纪次数"
                chart.x_axis.title = "班级"
                values = Reference(ws4, min_col=2, min_row=3, max_row=row-1)
                cats = Reference(ws4, min_col=1, min_row=4, max_row=row-1)
                chart.add_data(values, titles_from_data=True)
                chart.set_categories(cats)
                ws4.add_chart(chart, f"D3")

        # ── Sheet 5: 消息/活动汇总 ──
        ws5 = wb.create_sheet("消息活动")
        ws5['A1'] = "消息与活动汇总"
        ws5['A1'].font = Font(size=14, bold=True)

        if "message_stats" in data:
            ms = data["message_stats"]
            ws5['A3'] = "消息总数"; ws5['B3'] = ms.get("total", 0)
            ws5['A4'] = "未读数"; ws5['B4'] = ms.get("unread", 0)

        if "activity_stats" in data:
            act = data["activity_stats"]
            ws5['A6'] = "活动总数"; ws5['B6'] = act.get("total", 0)
            ws5['A7'] = "参与人次"; ws5['B7'] = act.get("participants", 0)

        # 设置列宽
        for ws in wb.worksheets:
            for col in range(1, 10):
                ws.column_dimensions[get_column_letter(col)].width = 18

        # 保存
        wb.save(file_path)

        # 更新数据库
        report.file_path = file_path
        safe_commit()

        return file_path
    except ImportError:
        return None
    except Exception as e:
        print(f"Excel生成失败: {e}")
        import traceback
        traceback.print_exc()
        return None

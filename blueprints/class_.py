"""班主任工作台 — 本班纪律/考勤/评分/重点关注/通知"""
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session, current_app
from sqlalchemy.orm import joinedload
from models import db, Student, Class, Task, TaskFeedback
from models import DisciplineRecord, RoutineScore, Attendance, LeaveRequest, ProblemStudent
from models import Message, Announcement, User
from blueprints.discipline_utils import check_escalation, send_discipline_notifications, deduct_quality_score
from blueprints.common import notify_parent, notify_class_teacher
from blueprints.audit_log import audit_log
from decorators import login_required, require_role, require_permission
from datetime import date, datetime, timezone, timedelta
from utils.db_utils import safe_commit
from utils import get_local_now

# 声呐事件总线（违纪实时广播）
from utils.sonar_bus import publish_discipline

class_bp = Blueprint("class", __name__, url_prefix="/class")


@class_bp.before_request
@login_required
@require_role("class_teacher", "ms_admin")
def check_role():
    pass


# ── 班主任工作台首页 ──
@class_bp.route("/")
def dashboard():
    class_id = session.get("class_id")
    if not class_id:
        flash("请先选择一个班级", "warning")
        return redirect(url_for("grade.dashboard"))
    class_obj = Class.query.get(class_id)
    today = date.today()
    # 直接 COUNT 代替全量 ORM 加载（节省 200ms）
    student_count = Student.query.filter_by(class_id=class_id, is_active=True).count()
    stats = {
        "student_count": student_count,
        "discipline_count": DisciplineRecord.query.filter_by(
            class_id=class_id, status="active").count(),
        "problem_count": ProblemStudent.query.filter_by(
            class_id=class_id, status="active").count(),
        "absent_today": Attendance.query.filter(
            Attendance.class_id == class_id,
            Attendance.record_date == today,
            Attendance.status.in_(["absent", "leave"])
        ).count(),
        "pending_tasks": Task.query.filter_by(
            target_type="class", target_id=class_id, status="pending").count(),
        "pending_leaves": LeaveRequest.query.filter_by(
            class_id=class_id, status="pending").count(),
    }
    problems = ProblemStudent.query.filter_by(class_id=class_id, status="active").options(
        joinedload(ProblemStudent.student)).all()
    return render_template("class_/dashboard.html", stats=stats, class_obj=class_obj,
                           problems=problems)


# ── 学生花名册（支持搜索/筛选） ──
@class_bp.route("/students")
def student_list():
    class_id = session.get("class_id")
    grade_id = session.get("grade_id")
    # Admin fallback: show first active class if no session class
    if not class_id:
        first = Class.query.filter_by(is_active=True).order_by(Class.name).first()
        if first:
            class_id = first.id
        else:
            return "系统中暂无班级数据", 400
    q = request.args.get("q", "").strip()
    gender = request.args.get("gender", "")
    tag = request.args.get("tag", "").strip()
    
    query = Student.query.filter_by(class_id=class_id, is_active=True)
    
    if q:
        like = f"%{q}%"
        query = query.filter(
            (Student.name.like(like)) | (Student.student_no.like(like))
        )
    if gender:
        query = query.filter_by(gender=gender)
    if tag:
        like_tag = f"%{tag}%"
        query = query.filter(Student.tags.like(like_tag))
    
    students = query.order_by(Student.student_no).all()
    # 如果请求 JSON，返回学生列表（供其他页面 AJAX 调用）
    if "application/json" in request.headers.get("Accept", ""):
        return jsonify({
            "students": [
                {"id": s.id, "name": s.name, "student_no": s.student_no,
                 "gender": s.gender, "class_id": s.class_id}
                for s in students
            ]
        })
    return render_template("class_/students.html", students=students)


@class_bp.route("/students/add", methods=["GET", "POST"])
@audit_log("add_student", "Student")
def add_student():
    """添加单个学生"""
    class_id = session.get("class_id")
    grade_id = session.get("grade_id")
    classes = Class.query.filter_by(is_active=True).order_by(Class.name).all()
    if not class_id and classes:
        class_id = classes[0].id
        grade_id = classes[0].grade_id
    
    if request.method == "POST":
        if session.get("role") == "ms_admin" and request.form.get("class_id"):
            class_id = int(request.form["class_id"])
            cls = next((c for c in classes if c.id == class_id), None)
            grade_id = cls.grade_id if cls else grade_id
        s = Student(
            name=request.form["name"],
            student_no=request.form["student_no"],
            class_id=class_id,
            grade_id=grade_id,
            gender=request.form.get("gender", "男"),
            id_card=request.form.get("id_card", ""),
            national_id=request.form.get("national_id", ""),
            ethnicity=request.form.get("ethnicity", "汉族"),
            birth_date=datetime.strptime(request.form["birth_date"], "%Y-%m-%d").date() if request.form.get("birth_date") else None,
            address=request.form.get("address", ""),
            parent1_name=request.form.get("parent1_name", ""),
            parent1_phone=request.form.get("parent1_phone", ""),
            parent1_relation=request.form.get("parent1_relation", ""),
            parent2_name=request.form.get("parent2_name", ""),
            parent2_phone=request.form.get("parent2_phone", ""),
            parent2_relation=request.form.get("parent2_relation", ""),
            primary_school=request.form.get("primary_school", ""),
            tags=request.form.get("tags", ""),
        )
        db.session.add(s)
        safe_commit()
        flash("学生已添加", "success")
        return redirect(url_for("class.student_list"))
    
    return render_template("class_/student_form.html", student=None, action="add")


@class_bp.route("/students/<int:sid>/edit", methods=["GET", "POST"])
@audit_log("edit_student", "Student")
def edit_student(sid):
    """编辑学生"""
    student = Student.query.get_or_404(sid)
    if student.class_id != session.get("class_id") and session.get("role") != "ms_admin":
        flash("无权操作", "danger")
        return redirect(url_for("class.student_list"))
    
    if request.method == "POST":
        student.name = request.form["name"]
        student.student_no = request.form["student_no"]
        student.gender = request.form.get("gender", "男")
        student.id_card = request.form.get("id_card", "")
        student.national_id = request.form.get("national_id", "")
        student.ethnicity = request.form.get("ethnicity", "汉族")
        student.birth_date = datetime.strptime(request.form["birth_date"], "%Y-%m-%d").date() if request.form.get("birth_date") else None
        student.address = request.form.get("address", "")
        student.parent1_name = request.form.get("parent1_name", "")
        student.parent1_phone = request.form.get("parent1_phone", "")
        student.parent1_relation = request.form.get("parent1_relation", "")
        student.parent2_name = request.form.get("parent2_name", "")
        student.parent2_phone = request.form.get("parent2_phone", "")
        student.parent2_relation = request.form.get("parent2_relation", "")
        student.primary_school = request.form.get("primary_school", "")
        student.tags = request.form.get("tags", "")
        safe_commit()
        flash("学生信息已更新", "success")
        return redirect(url_for("class.student_detail", sid=sid))
    
    return render_template("class_/student_form.html", student=student, action="edit")


@class_bp.route("/students/<int:sid>/delete", methods=["POST"])
@audit_log("delete_student", "Student")
def delete_student(sid):
    """软删除学生"""
    student = Student.query.get_or_404(sid)
    if student.class_id != session.get("class_id") and session.get("role") != "ms_admin":
        flash("无权操作", "danger")
        return redirect(url_for("class.student_list"))
    student.is_active = False
    safe_commit()
    flash("学生已删除", "success")
    return redirect(url_for("class.student_list"))


@class_bp.route("/students/import", methods=["POST"])
def import_students():
    """Excel批量导入学生"""
    from io import BytesIO
    from openpyxl import load_workbook
    
    f = request.files.get("file")
    if not f or not f.filename:
        flash("请选择文件", "danger")
        return redirect(url_for("class.student_list"))
    
    class_id = session.get("class_id")
    grade_id = session.get("grade_id")
    if not class_id:
        first = Class.query.filter_by(is_active=True).order_by(Class.name).first()
        if first:
            class_id = first.id
            grade_id = first.grade_id
        else:
            flash("系统中暂无班级", "danger")
            return redirect(url_for("class.student_list"))
    wb = load_workbook(BytesIO(f.read()))
    ws = wb.active
    
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        no = str(row[0]).strip()
        name = str(row[1]).strip() if row[1] else ""
        if not name:
            continue
        if Student.query.filter_by(student_no=no).first():
            continue  # 跳过重复学号
        s = Student(
            student_no=no,
            name=name,
            class_id=class_id,
            grade_id=grade_id,
            gender=str(row[2]).strip() if row[2] else "男",
            id_card=str(row[3]).strip() if row[3] else "",
            national_id=str(row[4]).strip() if row[4] else "",
            parent1_name=str(row[5]).strip() if row[5] else "",
            parent1_phone=str(row[6]).strip() if row[6] else "",
        )
        db.session.add(s)
        count += 1
    
    safe_commit()
    flash(f"成功导入 {count} 名学生", "success")
    return redirect(url_for("class.student_list"))


@class_bp.route("/students/template")
def download_template():
    """下载Excel导入模板"""
    from openpyxl import Workbook
    from io import BytesIO
    
    wb = Workbook()
    ws = wb.active
    ws.title = "学生导入模板"
    ws.append(["学号*", "姓名*", "性别", "身份证号", "全国学籍号", "家长1姓名", "家长1电话"])
    ws.append(["2024001", "张三", "男", "430XXXXXXX", "GXXXXXXXX", "张三父", "13800000000"])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from flask import send_file
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name="student_template.xlsx")


@class_bp.route("/students/export")
def export_students():
    """导出本班学生到Excel"""
    from openpyxl import Workbook
    from io import BytesIO
    
    class_id = session.get("class_id")
    students = Student.query.filter_by(class_id=class_id, is_active=True).order_by(Student.student_no).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "学生名册"
    ws.append(["学号", "姓名", "性别", "家长1", "电话1", "家长2", "电话2", "标签"])
    
    for s in students:
        ws.append([
            s.student_no, s.name, s.gender,
            s.parent1_name, s.parent1_phone,
            s.parent2_name, s.parent2_phone,
            s.tags
        ])
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    from flask import send_file
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=f"students_{class_id}.xlsx")


@class_bp.route("/students/<int:sid>")
def student_detail(sid):
    student = Student.query.get_or_404(sid)
    if student.class_id != session.get("class_id") and session.get("role") != "ms_admin":
        flash("无权查看", "danger")
        return redirect(url_for("class.student_list"))
    disciplines = DisciplineRecord.query.filter_by(student_id=sid).order_by(
        DisciplineRecord.created_at.desc()).all()
    attendances = Attendance.query.filter_by(student_id=sid).order_by(
        Attendance.record_date.desc()).limit(30).all()
    return render_template("class_/student_detail.html", student=student,
                           disciplines=disciplines, attendances=attendances)


# ── 纪律记录 ──
@class_bp.route("/discipline")
def discipline_list():
    class_id = session.get("class_id")
    page = request.args.get("page", 1, type=int)
    records = DisciplineRecord.query.filter_by(class_id=class_id).options(
        joinedload(DisciplineRecord.student)
    ).order_by(
        DisciplineRecord.created_at.desc()).paginate(page=page, per_page=20)
    students = Student.query.filter_by(class_id=class_id, is_active=True).all()
    return render_template("class_/discipline.html", records=records, students=students)


@class_bp.route("/discipline/add", methods=["POST"])
@audit_log("add_discipline", "DisciplineRecord")
def add_discipline():
    student_id = request.form.get("student_id", type=int)
    student = Student.query.get_or_404(student_id)
    # 班主任只能添加本班学生的违纪记录
    if session.get("role") == "class_teacher" and student.class_id != session.get("class_id"):
        flash("无权操作该学生", "danger")
        return redirect(url_for("class.discipline_list"))
    record = DisciplineRecord(
        student_id=student.id,
        class_id=student.class_id,
        grade_id=student.grade_id,
        type=request.form.get("type", "minor"),
        category=request.form.get("category", ""),
        description=request.form.get("description", ""),
        action_taken=request.form.get("action_taken", ""),
        points=request.form.get("points", 0, type=int),
        created_by=session.get("user_id"),
        verify_status="VERIFIED",
    )
    db.session.add(record)
    # 积分累计自动升级 + 通知（与违纪记录同一个事务）
    check_escalation(student, session.get("user_id"))
    send_discipline_notifications(record, student)
    deduct_quality_score(record, student, session.get("user_id"))

    safe_commit()
    # ── 声呐广播：违纪实时推送 ──
    try:
        publish_discipline(record, session.get("display_name", ""))
    except Exception:
        pass
    flash("违纪记录已添加", "success")
    return redirect(url_for("class.discipline_list"))


# ── 缺勤批量转违纪（终极缝合：副作用队列 + 事务隔离 + 全联动保证） ──
@class_bp.route("/discipline/from-attendance", methods=["GET", "POST"])
@require_permission("manage_discipline")
def discipline_from_attendance():
    class_id = session.get("class_id")
    today = date.today()

    # GET: 显示今日缺勤列表（保持原入口，向后兼容）
    if request.method == "GET":
        absent_records = Attendance.query.filter(
            Attendance.class_id == class_id,
            Attendance.record_date == today,
            Attendance.status.in_(["absent", "late", "leave"]),
        ).order_by(Attendance.student_id).all()

        student_ids = [r.student_id for r in absent_records]
        students_map = {}
        if student_ids:
            for s in Student.query.filter(Student.id.in_(student_ids)).all():
                students_map[s.id] = s

        return render_template("class_/discipline_from_attendance.html",
                               records=absent_records, students_map=students_map, today=today)

    # POST: 执行批量转化（副作用队列模式）
    selected_aids = request.form.getlist("attendance_ids")
    if not selected_aids:
        flash("请选择需要处理的考勤记录", "warning")
        return redirect(url_for("class.discipline_list"))

    # 输入验证：防 500 崩溃
    attendance_ids = [int(aid) for aid in selected_aids if aid.isdigit()]
    if not attendance_ids:
        flash("未找到有效的考勤记录", "warning")
        return redirect(url_for("class.discipline_list"))

    # joinedload 消灭 N+1
    attendance_records = Attendance.query.options(
        joinedload(Attendance.student)
    ).filter(Attendance.id.in_(attendance_ids)).all()

    form_points = request.form.get("points", type=int) or 1
    action_taken = request.form.get("action_taken", "")

    # 副作用队列：先建记录，主事务落盘成功后再执行外部联动
    side_effects_queue = []
    new_records_count = 0

    for att in attendance_records:
        student = att.student
        if not student:
            continue

        # 班主任只能操作本班学生
        if session.get("role") == "class_teacher" and student.class_id != class_id:
            continue

        # 根据考勤状态自动判定违纪类型
        record = DisciplineRecord(
            student_id=student.id,
            class_id=student.class_id,
            grade_id=student.grade_id,
            type="absent" if att.status == "absent" else "minor",
            category="考勤联动",
            description=f"由考勤记录自动联动生成：状态为 [{att.status}]，日期为 {att.record_date}",
            action_taken=action_taken or "待班主任处理",
            points=form_points,
            created_by=session.get("user_id"),
            verify_status="VERIFIED",
            created_at=get_local_now(),
        )
        db.session.add(record)
        db.session.flush()  # 获取 record.id
        new_records_count += 1

        # 将待触发的联动操作压入队列，暂不执行
        side_effects_queue.append((record, student))

    if new_records_count == 0:
        flash("未发现可处理的有效考勤记录", "info")
        return redirect(url_for("class.discipline_list"))

    # 主事务落盘
    safe_commit()

    # 主记录已保存，执行副作用队列（外部通知 + 后续扣分）
    for record, student in side_effects_queue:
        try:
            check_escalation(student, session.get("user_id"))
            send_discipline_notifications(record, student)
            deduct_quality_score(record, student, session.get("user_id"))
        except Exception as e:
            current_app.logger.error(
                f"联动副作用触发失败 (RecordID: {record.id}, Student: {student.name}): {e}"
            )

    # 副作用落盘（失败不影响主记录结果）
    try:
        safe_commit()
    except RuntimeError as e:
        current_app.logger.critical(
            f"副作用事务提交失败 (已生成 {new_records_count} 条违纪记录): {e}"
        )

    flash(f"已成功将 {new_records_count} 条考勤记录转化为违纪处分", "success")

    return redirect(url_for("class.discipline_list"))


@class_bp.route("/discipline/<int:rid>/edit", methods=["POST"])
def edit_discipline(rid):
    """编辑违纪记录"""
    record = DisciplineRecord.query.get_or_404(rid)
    if record.class_id != session.get("class_id") and session.get("role") != "ms_admin":
        flash("无权操作", "danger")
        return redirect(url_for("class.discipline_list"))
    record.type = request.form.get("type", record.type)
    record.category = request.form.get("category", record.category)
    record.description = request.form.get("description", record.description)
    record.action_taken = request.form.get("action_taken", record.action_taken)
    record.points = request.form.get("points", record.points, type=int)
    safe_commit()
    flash("违纪记录已更新", "success")
    return redirect(url_for("class.discipline_list"))


@class_bp.route("/discipline/<int:rid>/delete", methods=["POST"])
@require_permission("manage_discipline")
def delete_discipline(rid):
    """删除违纪记录（需 manage_discipline 权限）"""
    record = DisciplineRecord.query.get_or_404(rid)
    # 越权防护强化：非德育处全局管理员，必须满足班级归属一致
    if session.get("role") != "ms_admin" and record.class_id != session.get("class_id"):
        flash("越权操作：您无权删除非本班学生的违纪记录", "danger")
        return redirect(url_for("class.discipline_list"))
    db.session.delete(record)
    safe_commit()
    flash("违纪记录已撤销/删除", "success")
    return redirect(url_for("class.discipline_list"))


@class_bp.route("/discipline/<int:rid>/resolve", methods=["POST"])
@require_permission("manage_discipline")
def resolve_discipline(rid):
    """标记违纪记录为已解决（需 manage_discipline 权限）"""
    record = DisciplineRecord.query.get_or_404(rid)
    # 越权防护强化
    if session.get("role") != "ms_admin" and record.class_id != session.get("class_id"):
        flash("越权操作：您无权处理该班级的违纪申诉", "danger")
        return redirect(url_for("class.discipline_list"))
    record.status = "resolved"
    record.resolved_at = get_local_now()  # 修复：废除 utcnow，统一本地时间
    safe_commit()
    flash("该违纪记录已被标记为已解决", "success")
    return redirect(url_for("class.discipline_list"))


# ── 批量违纪添加 ──
@class_bp.route("/discipline/batch", methods=["GET", "POST"])
@login_required
@require_role("class_teacher")
def batch_discipline():
    """批量添加违纪记录"""
    class_id = session.get("class_id")
    students = Student.query.filter_by(class_id=class_id, is_active=True).order_by(
        Student.student_no).all()

    if request.method == "POST":
        selected_ids = request.form.getlist("student_ids")
        if not selected_ids:
            flash("请选择至少一名学生", "warning")
            return redirect(url_for("class.batch_discipline"))

        d_type = request.form.get("type", "minor")
        category = request.form.get("category", "")
        description = request.form.get("description", "")
        action_taken = request.form.get("action_taken", "")
        points = request.form.get("points", 0, type=int)

        # 修复: 批量加载学生，消灭 N+1 查询
        sids = [int(x) for x in selected_ids]
        sel_students = Student.query.filter(Student.id.in_(sids)).all()
        students_map = {s.id: s for s in sel_students}

        count = 0
        for sid in sids:
            student = students_map.get(sid)
            if not student:
                continue
            record = DisciplineRecord(
                student_id=student.id,
                class_id=class_id,
                grade_id=student.grade_id,
                type=d_type,
                category=category,
                description=f"[批量] {description}",
                action_taken=action_taken or "待班主任处理",
                points=points,
                created_by=session.get("user_id"),
                verify_status="VERIFIED",
            )
            db.session.add(record)
            # 升级检查 + 通知 + 素质扣分
            check_escalation(student, session.get("user_id"))
            send_discipline_notifications(record, student)
            deduct_quality_score(record, student, session.get("user_id"))
            count += 1

        safe_commit()
        flash(f"已批量添加 {count} 条违纪记录", "success")
        return redirect(url_for("class.discipline_list"))

    return render_template("class_/discipline_batch.html", students=students)


# ── 考勤 ──
@class_bp.route("/attendance", methods=["GET", "POST"])
def attendance_page():
    class_id = session.get("class_id")
    grade_id = session.get("grade_id")
    if not class_id:
        first = Class.query.filter_by(is_active=True).order_by(Class.name).first()
        if first:
            class_id = first.id
            grade_id = first.grade_id
    today = date.today()
    if request.method == "POST":
        from_user_id = session.get("user_id")
        absent_students = []  # 缺勤学生
        late_students = []    # 迟到学生

        for key, value in request.form.items():
            if key.startswith("status_"):
                sid = int(key.replace("status_", ""))
                Attendance.query.filter_by(
                    student_id=sid, record_date=today
                ).delete()
                db.session.add(Attendance(
                    student_id=sid,
                    class_id=class_id,
                    grade_id=grade_id,
                    status=value,
                    record_date=today,
                ))

                # 记录异常
                if value == "absent":
                    absent_students.append(sid)
                elif value == "late":
                    late_students.append(sid)

        safe_commit()

        # ── 考勤异常通知：缺勤/迟到 → 班主任 + 家长 ──
        _notify_attendance_anomalies(
            absent_students, late_students, today, from_user_id
        )

        flash("考勤已保存", "success")
        return redirect(url_for("class.attendance_page"))
    students = Student.query.filter_by(class_id=class_id, is_active=True).all()
    today_records = {
        r.student_id: r.status
        for r in Attendance.query.filter_by(class_id=class_id, record_date=today).all()
    }
    return render_template("class_/attendance.html", students=students,
                           today_records=today_records, today=today)


@class_bp.route("/attendance/history")
def attendance_history():
    """考勤历史记录"""
    class_id = session.get("class_id")
    if not class_id:
        first = Class.query.filter_by(is_active=True).order_by(Class.name).first()
        if first:
            class_id = first.id
        else:
            return "暂无班级", 400
    # 默认显示最近30天
    days = request.args.get("days", 30, type=int)
    from datetime import timedelta
    end_date = date.today()
    start_date = end_date - timedelta(days=days - 1)
    
    records = (Attendance.query
               .filter(Attendance.class_id == class_id,
                       Attendance.record_date >= start_date,
                       Attendance.record_date <= end_date)
               .order_by(Attendance.record_date.desc())
               .all())
    
    # 按日期分组
    from collections import OrderedDict
    records_by_date = OrderedDict()
    for r in records:
        d = r.record_date
        if d not in records_by_date:
            records_by_date[d] = []
        records_by_date[d].append(r)
    
    # 统计
    total = len(records)
    stats = {"present": 0, "late": 0, "early": 0, "absent": 0, "leave": 0}
    for r in records:
        stats[r.status] = stats.get(r.status, 0) + 1
    
    students = {s.id: s for s in Student.query.filter_by(class_id=class_id, is_active=True).all()}
    
    return render_template("class_/attendance_history.html",
                           records_by_date=records_by_date, stats=stats,
                           days=days, today=date.today(), students=students, total=total)


# ── 请假审批（班主任初审） ──
@class_bp.route("/leaves")
def leave_list():
    class_id = session.get("class_id")
    leaves = LeaveRequest.query.filter_by(class_id=class_id).options(
        joinedload(LeaveRequest.student)
    ).order_by(
        LeaveRequest.created_at.desc()).limit(30).all()
    return render_template("class_/leaves.html", leaves=leaves)


@class_bp.route("/leaves/<int:lid>/approve", methods=["POST"])
@audit_log("approve_leave", "LeaveRequest")
def approve_leave(lid):
    leave = LeaveRequest.query.get_or_404(lid)
    if leave.class_id != session.get("class_id"):
        flash("无权审批", "danger")
        return redirect(url_for("class.leave_list"))
    action = request.form.get("action")
    if action == "approve":
        leave.status = "class_approved"
        leave.class_approved_by = session.get("user_id")
        leave.class_approved_at = date.today()
    else:
        leave.status = "rejected"
    safe_commit()

    # 通知家长
    student = Student.query.get(leave.student_id)
    if student:
        from_user_id = session.get("user_id")
        action_label = "通过（班主任已审批）" if leave.status == "class_approved" else "被驳回"
        notify_parent(
            student,
            title=f"请假审批结果 — {student.name}",
            content=f"您孩子 {student.name} 的请假申请已{action_label}。\n"
                    f"请假时间：{leave.start_date} ~ {leave.end_date}\n"
                    f"请假原因：{leave.reason}",
            from_user_id=from_user_id,
        )

    flash("审批完成", "success")
    return redirect(url_for("class.leave_list"))


# ── 任务反馈 ──
@class_bp.route("/tasks")
def task_list():
    class_id = session.get("class_id")
    status = request.args.get("status", "")
    q = Task.query.filter_by(target_type="class", target_id=class_id)
    if status:
        q = q.filter_by(status=status)
    tasks = q.order_by(Task.created_at.desc()).all()
    status_counts = {
        "pending": Task.query.filter_by(target_type="class", target_id=class_id, status="pending").count(),
        "done": Task.query.filter_by(target_type="class", target_id=class_id, status="done").count(),
        "closed": Task.query.filter_by(target_type="class", target_id=class_id, status="closed").count(),
    }
    # 修复: 批量加载所有反馈，消灭 N+1 查询
    task_ids = [t.id for t in tasks]
    feedbacks_by_task = {}
    if task_ids:
        all_fbs = (TaskFeedback.query
                   .filter(TaskFeedback.task_id.in_(task_ids))
                   .order_by(TaskFeedback.created_at.asc()).all())
        for fb in all_fbs:
            feedbacks_by_task.setdefault(fb.task_id, []).append(fb)
    # 确保每个任务都有列表
    for t in tasks:
        if t.id not in feedbacks_by_task:
            feedbacks_by_task[t.id] = []
    return render_template("class_/tasks.html", tasks=tasks, status=status,
                           status_counts=status_counts, feedbacks_by_task=feedbacks_by_task)


@class_bp.route("/tasks/<int:tid>/feedback", methods=["POST"])
def task_feedback(tid):
    """提交任务反馈"""
    # 空内容校验：防 500 崩溃
    content = request.form.get("content", "").strip()
    if not content:
        flash("反馈内容不能为空", "warning")
        return redirect(url_for("class.task_list"))

    fb = TaskFeedback(
        task_id=tid,
        user_id=session.get("user_id"),
        content=content,
        created_at=get_local_now(),
    )
    db.session.add(fb)

    # 任务标记：仅对未处理任务修改状态
    task = Task.query.get(tid)
    if task and task.status in ["pending", "assigned"]:
        task.status = "done"
        task.finished_at = get_local_now()

    safe_commit()
    flash("任务反馈提交成功", "success")
    return redirect(url_for("class.task_list"))


# ── 通知家长（发消息） ──
@class_bp.route("/notify", methods=["GET", "POST"])
def notify_parent_page():
    class_id = session.get("class_id")
    if request.method == "POST":
        student_ids = request.form.getlist("student_ids")
        # 修复: 批量加载学生和家长，消灭 N+1 查询
        sids_int = [int(x) for x in student_ids]
        students_map = {s.id: s for s in Student.query.filter(Student.id.in_(sids_int)).all()}
        parents_map = {}
        all_parents = User.query.filter(User.bound_student_id.in_(sids_int), User.role == "parent").all()
        for p in all_parents:
            parents_map.setdefault(p.bound_student_id, []).append(p)

        for sid in sids_int:
            student = students_map.get(sid)
            if not student:
                continue
            # 找绑定家长
            parents = parents_map.get(sid, [])
            for p in parents:
                msg = Message(
                    from_user_id=session.get("user_id"),
                    to_user_id=p.id,
                    title=request.form.get("title", "班主任通知"),
                    content=request.form.get("content", ""),
                )
                db.session.add(msg)
        safe_commit()
        flash("通知已发送", "success")
        return redirect(url_for("class.dashboard"))
    students = Student.query.filter_by(class_id=class_id, is_active=True).all()
    return render_template("class_/notify.html", students=students)


# ── 考勤异常通知 ─────────────────────────────
def _notify_attendance_anomalies(absent_ids, late_ids, record_date, from_user_id=None):
    """考勤异常时通知家长和班主任

    缺勤：通知班主任 + 家长（需关注）
    迟到：仅通知家长（提醒）
    """
    if not absent_ids and not late_ids:
        return

    date_str = record_date.strftime("%Y-%m-%d")
    all_student_ids = set(absent_ids) | set(late_ids)
    students = {s.id: s for s in Student.query.filter(
        Student.id.in_(all_student_ids)
    ).all()} if all_student_ids else {}

    for sid in absent_ids:
        stu = students.get(sid)
        if not stu:
            continue
        title = f"⚠️ 缺勤通知 — {stu.name}"
        content = f"学生 {stu.name} 于 {date_str} 被记录为缺勤。请及时关注学生动态，了解缺勤原因。"
        # 通知班主任
        notify_class_teacher(stu, title, content, from_user_id=from_user_id)
        # 通知家长
        notify_parent(stu, title, content, from_user_id=from_user_id)

    for sid in late_ids:
        stu = students.get(sid)
        if not stu:
            continue
        title = f"迟到提醒 — {stu.name}"
        content = f"学生 {stu.name} 于 {date_str} 被记录为迟到。请提醒孩子按时到校。"
        # 通知家长
        notify_parent(stu, title, content, from_user_id=from_user_id)

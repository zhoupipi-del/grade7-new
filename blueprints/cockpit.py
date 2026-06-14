"""数据驾驶舱 — 交互式全景仪表盘 + 德育声呐实时战情大屏"""
from flask import (Blueprint, render_template, request, jsonify, session,
                   send_file, Response, stream_with_context, url_for)
from models import (db, Student, Class, Grade, User, Subject, Exam, Score,
                    DisciplineRecord, Attendance, RoutineScore, WingsScore,
                    Notice, NoticeReceipt, HomeVisit, LeaveRequest,
                    MentalHealthAssessment, RiskRecord, Activity, ActivityRegistration,
                    ParentMeeting, ParentMeetingSignin)
from decorators import login_required, require_role, scope_query
from datetime import date, timedelta, datetime
from sqlalchemy import func, case, text
from utils.db_utils import safe_commit
import json, time, io

cockpit_bp = Blueprint("cockpit", __name__, template_folder="../templates/cockpit")

# ── 德育声呐全局广播频道 ──
SONAR_CHANNEL = "sse:sonar:global"
MAX_SONAR_STREAM = 55  # 秒
MAX_HISTORY = 50  # 内存中保留的最大事件数

# 内存事件回放队列（跨请求共享）
_sonar_history = []


@cockpit_bp.route("/")
@login_required
def index():
    """驾驶舱首页"""
    grade_id = session.get("grade_id")
    grades = Grade.query.order_by(Grade.sort_order).all()
    exams = Exam.query.order_by(Exam.exam_date.desc()).all()
    return render_template("cockpit/index.html",
                           grades=grades, exams=exams,
                           current_grade_id=grade_id)


@cockpit_bp.route("/data")
@login_required
def data_api():
    """全景数据 API — 支持筛选参数（已优化 N+1 查询）"""
    grade_id = request.args.get("grade_id", type=int) or session.get("grade_id")
    if not grade_id:
        g = Grade.query.order_by(Grade.sort_order).first()
        grade_id = g.id if g else 1

    exam_id = request.args.get("exam_id", type=int)
    days = request.args.get("days", 30, type=int)
    since = date.today() - timedelta(days=days)

    # ── 基础统计 ──
    total_students = Student.query.filter_by(grade_id=grade_id, is_active=True).count()
    total_classes = Class.query.filter_by(grade_id=grade_id, is_active=True).count()
    teachers = User.query.filter_by(grade_id=grade_id, role="class_teacher").count()
    grade_leaders = User.query.filter_by(grade_id=grade_id, role="grade_leader").count()
    total_teachers = teachers + grade_leaders

    # ── 成绩统计（优化：纯 SQL 聚合，零 ORM 对象加载）──
    score_stats = {}
    
    subject_count = Subject.query.count()
    full_total = subject_count * 100
    
    if not exam_id:
        latest_exam = Exam.query.filter_by(grade_id=grade_id).order_by(Exam.exam_date.desc()).first()
        if latest_exam:
            exam_id = latest_exam.id
    
    if exam_id:
        row = db.session.execute(text("""
            SELECT
                COUNT(*) AS total_participants,
                ROUND(AVG(total), 1) AS avg_score,
                MAX(total) AS max_score,
                MIN(total) AS min_score,
                SUM(CASE WHEN total >= :pass_cut THEN 1 ELSE 0 END) AS pass_count,
                SUM(CASE WHEN total >= :excel_cut THEN 1 ELSE 0 END) AS excellent_count
            FROM (
                SELECT student_id, SUM(score) AS total
                FROM scores
                WHERE exam_id = :eid AND grade_id = :gid
                GROUP BY student_id
            ) t
        """), {"eid": exam_id, "gid": grade_id, "pass_cut": full_total * 0.6, "excel_cut": full_total * 0.85}).fetchone()
        
        if row:
            tp = row.total_participants
            score_stats = {
                "avg_score": float(row.avg_score) if row.avg_score else 0,
                "max_score": int(row.max_score) if row.max_score else 0,
                "min_score": int(row.min_score) if row.min_score else 0,
                "pass_count": row.pass_count or 0,
                "excellent_count": row.excellent_count or 0,
                "total_participants": tp,
                "pass_rate": round(row.pass_count / tp * 100, 1) if tp else 0,
                "excellent_rate": round(row.excellent_count / tp * 100, 1) if tp else 0,
            }

    # ── 德育统计（优化：一次聚合查询）──
    disc_query = DisciplineRecord.query.filter(
        DisciplineRecord.grade_id == grade_id,
        DisciplineRecord.created_at >= since,
    )
    discipline_count = disc_query.count()
    
    # 一次查询按类型和分类聚合
    discipline_by = db.session.query(
        DisciplineRecord.type,
        DisciplineRecord.category,
        func.count(DisciplineRecord.id)
    ).filter(
        DisciplineRecord.grade_id == grade_id,
        DisciplineRecord.created_at >= since,
    ).group_by(DisciplineRecord.type, DisciplineRecord.category).all()
    
    disc_levels = {"warning": 0, "minor": 0, "major": 0, "serious": 0}
    discipline_by_category = []
    cat_map = {}
    for dtype, cat, cnt in discipline_by:
        disc_levels[dtype] = cnt + disc_levels.get(dtype, 0)
        cat_key = cat or "未分类"
        cat_map[cat_key] = cat_map.get(cat_key, 0) + cnt
    discipline_by_category = [{"name": k, "count": v} for k, v in cat_map.items()]

    # ── 考勤统计（优化：一次聚合查询）──
    att_stats = db.session.query(
        Attendance.status,
        func.count(Attendance.id)
    ).filter(
        Attendance.grade_id == grade_id,
        Attendance.record_date >= since,
    ).group_by(Attendance.status).all()
    
    att_status = {"present": 0, "late": 0, "early": 0, "absent": 0, "leave": 0}
    total_att = 0
    for status, cnt in att_stats:
        att_status[status] = cnt
        total_att += cnt
    attendance_rate = round(att_status["present"] / total_att * 100, 1) if total_att > 0 else 0

    # 请假统计
    leave_count = LeaveRequest.query.filter(
        LeaveRequest.grade_id == grade_id,
        LeaveRequest.created_at >= since,
    ).count()
    leave_approved = LeaveRequest.query.filter(
        LeaveRequest.grade_id == grade_id,
        LeaveRequest.created_at >= since,
        LeaveRequest.status == "approved",
    ).count()

    # ── 通知统计 ──
    notice_count = Notice.query.filter(
        Notice.grade_id == grade_id,
        Notice.created_at >= since,
    ).count()
    total_notices = Notice.query.filter_by(grade_id=grade_id).count()

    # 阅读率（优化：一次 JOIN 替代两次查询）
    notice_read_rate = 0
    if total_notices > 0 and total_students > 0:
        total_receipts_needed = total_notices * total_students
        read_count = db.session.execute(text("""
            SELECT COUNT(nr.id)
            FROM notice_receipts nr
            JOIN notices n ON nr.notice_id = n.id
            WHERE n.grade_id = :gid AND nr.status IN ('read', 'signed')
        """), {"gid": grade_id}).scalar()
        notice_read_rate = round(int(read_count or 0) / total_receipts_needed * 100, 1) if read_count else 0

    # ── 家访统计（优化：一次聚合查询）──
    visit_count = HomeVisit.query.filter(
        HomeVisit.grade_id == grade_id,
        HomeVisit.visit_date >= since,
    ).count()

    visit_by_type = db.session.query(
        HomeVisit.visit_type,
        func.count(HomeVisit.id)
    ).filter(
        HomeVisit.grade_id == grade_id,
        HomeVisit.visit_date >= since,
    ).group_by(HomeVisit.visit_type).all()

    # ── 心理健康风险分布 ──
    mh_stats = db.session.query(
        MentalHealthAssessment.risk_level,
        func.count(MentalHealthAssessment.id)
    ).filter_by(grade_id=grade_id).group_by(MentalHealthAssessment.risk_level).all()
    mh_risk = {"high": 0, "medium": 0, "low": 0}
    for level, cnt in mh_stats:
        mh_risk[level] = cnt

    # ── AI预警统计（最近扫描）──
    latest_scan = db.session.query(func.max(RiskRecord.scan_date)).filter_by(grade_id=grade_id).scalar()
    risk_stats = {"red": 0, "yellow": 0, "green": 0}
    if latest_scan:
        risk_data = db.session.query(
            RiskRecord.risk_level,
            func.count(RiskRecord.id)
        ).filter_by(grade_id=grade_id, scan_date=latest_scan).group_by(RiskRecord.risk_level).all()
        for level, cnt in risk_data:
            risk_stats[level] = cnt
    risk_scan_date = latest_scan.strftime("%Y-%m-%d") if latest_scan else None

    # ── 活动参与度（优化：一次 JOIN 替代两次查询）──
    total_activities = Activity.query.filter(
        Activity.grade_id.in_([grade_id, None]),
        Activity.status.in_(["completed", "ongoing"]),
    ).count()
    
    activity_reg_count = 0
    if total_activities > 0:
        activity_reg_count = db.session.execute(text("""
            SELECT COUNT(ar.id)
            FROM activity_registrations ar
            JOIN activities a ON ar.activity_id = a.id
            WHERE (a.grade_id = :gid_self OR a.grade_id IS NULL)
              AND a.status IN ('completed', 'ongoing')
              AND ar.status = 'confirmed'
        """), {"gid_self": grade_id}).scalar() or 0
        activity_reg_count = int(activity_reg_count)

    # ── 家长会参与率（优化：一次 JOIN 替代两次查询）──
    pm_total = ParentMeeting.query.filter_by(grade_id=grade_id).count()
    pm_signin_count = 0
    pm_rate = 0
    if pm_total > 0:
        pm_signin_count = db.session.execute(text("""
            SELECT COUNT(pms.id)
            FROM parent_meeting_signins pms
            JOIN parent_meetings pm ON pms.meeting_id = pm.id
            WHERE pm.grade_id = :gid
        """), {"gid": grade_id}).scalar() or 0
        pm_rate = round(int(pm_signin_count) / (pm_total * 30) * 100, 1)

    # ── 五翼均分（优化：一次聚合查询）──
    wing_dimensions = db.session.query(
        WingsScore.dimension,
        func.avg(WingsScore.score)
    ).filter(
        WingsScore.grade_id == grade_id,
    ).group_by(WingsScore.dimension).all()
    
    wing_avg_result = db.session.query(
        func.avg(WingsScore.score)
    ).filter(
        WingsScore.grade_id == grade_id,
    ).scalar()
    wing_avg = round(float(wing_avg_result), 1) if wing_avg_result else 0

    # ── 班级对比数据（优化：SQL 聚合，零 Python 循环）──
    classes = Class.query.filter_by(grade_id=grade_id, is_active=True).order_by(Class.name).all()
    
    class_score_data = []
    if exam_id:
        class_rows = db.session.execute(text("""
            SELECT
                class_id,
                COUNT(DISTINCT student_id) AS cnt,
                ROUND(AVG(total), 1) AS avg_score
            FROM (
                SELECT student_id, class_id, SUM(score) AS total
                FROM scores
                WHERE exam_id = :eid AND grade_id = :gid
                GROUP BY student_id, class_id
            ) t
            GROUP BY class_id
        """), {"eid": exam_id, "gid": grade_id}).fetchall()
        
        class_map = {r.class_id: (float(r.avg_score) if r.avg_score else 0, r.cnt) for r in class_rows}
        for cls in classes:
            a, c = class_map.get(cls.id, (0, 0))
            class_score_data.append({"name": cls.name, "score": a, "count": c})
    else:
        for cls in classes:
            class_score_data.append({"name": cls.name, "score": 0, "count": 0})

    # 各班违纪数（优化：一次聚合查询）
    class_disc_data = []
    disc_by_class = dict(db.session.query(
        DisciplineRecord.class_id,
        func.count(DisciplineRecord.id)
    ).filter(
        DisciplineRecord.grade_id == grade_id,
        DisciplineRecord.created_at >= since,
    ).group_by(DisciplineRecord.class_id).all())
    
    for cls in classes:
        class_disc_data.append({"name": cls.name, "count": disc_by_class.get(cls.id, 0)})

    # 各班出勤率（优化：一次聚合查询）
    class_att_data = []
    att_by_class = {}
    for cls_id, status, cnt in db.session.query(
        Attendance.class_id,
        Attendance.status,
        func.count(Attendance.id)
    ).filter(
        Attendance.grade_id == grade_id,
        Attendance.record_date >= since,
    ).group_by(Attendance.class_id, Attendance.status).all():
        if cls_id not in att_by_class:
            att_by_class[cls_id] = {"total": 0, "present": 0}
        att_by_class[cls_id]["total"] += cnt
        if status == "present":
            att_by_class[cls_id]["present"] += cnt
    
    for cls in classes:
        att = att_by_class.get(cls.id, {"total": 0, "present": 0})
        rate = round(att["present"] / att["total"] * 100, 1) if att["total"] > 0 else 0
        class_att_data.append({"name": cls.name, "rate": rate})

    # ── 趋势数据（优化：预加载 + 内存聚合，消灭 90 次查询！）──
    trend_dates = []
    trend_discipline = []
    trend_attendance = []
    trend_routine = []

    # 一次查询拿到最近 days 天的所有违纪记录
    disc_since = DisciplineRecord.query.filter(
        DisciplineRecord.grade_id == grade_id,
        DisciplineRecord.created_at >= since,
    ).all()
    disc_by_date = {}
    for d in disc_since:
        d_key = d.created_at.date() if d.created_at else date.today()
        disc_by_date[d_key] = disc_by_date.get(d_key, 0) + 1

    # 一次查询拿到最近 days 天的所有考勤记录
    att_since = Attendance.query.filter(
        Attendance.grade_id == grade_id,
        Attendance.record_date >= since,
    ).all()
    att_by_date = {}
    for a in att_since:
        d_key = a.record_date
        if d_key not in att_by_date:
            att_by_date[d_key] = {"total": 0, "present": 0}
        att_by_date[d_key]["total"] += 1
        if a.status == "present":
            att_by_date[d_key]["present"] += 1

    # 一次查询拿到最近 days 天的所有常规评分
    routine_since = RoutineScore.query.filter(
        RoutineScore.grade_id == grade_id,
        RoutineScore.created_at >= since,
    ).all()
    routine_by_date = {}
    for r in routine_since:
        d_key = r.created_at.date() if r.created_at else date.today()
        if d_key not in routine_by_date:
            routine_by_date[d_key] = []
        routine_by_date[d_key].append(float(r.score) if r.score else 0)

    # 现在循环日期，从内存字典中取数据（0 次数据库查询！）
    for i in range(days - 1, -1, -1):
        d = date.today() - timedelta(days=i)
        trend_dates.append(d.strftime("%m-%d"))
        trend_discipline.append(disc_by_date.get(d, 0))
        
        att = att_by_date.get(d, {"total": 0, "present": 0})
        trend_attendance.append(round(att["present"] / att["total"] * 100, 1) if att["total"] > 0 else None)
        
        rs = routine_by_date.get(d, [])
        trend_routine.append(round(sum(rs) / len(rs), 1) if rs else None)

    # ── 跨考试趋势（优化：一次查询 + 内存聚合）──
    all_exams = Exam.query.filter_by(grade_id=grade_id).order_by(Exam.exam_date.asc()).all()
    exam_trend_labels = []
    exam_trend_avgs = []
    exam_trend_pass = []
    exam_trend_excel = []

    if all_exams:
        # 一次查询拿到所有考试的所有成绩
        exam_ids = [e.id for e in all_exams]
        all_scores_by_exam = {}
        for s in Score.query.filter(
            Score.exam_id.in_(exam_ids),
            Score.grade_id == grade_id,
        ).all():
            if s.exam_id not in all_scores_by_exam:
                all_scores_by_exam[s.exam_id] = {}
            all_scores_by_exam[s.exam_id][s.student_id] = all_scores_by_exam[s.exam_id].get(s.student_id, 0) + s.score
        
        subject_count = Subject.query.count()
        full_total = subject_count * 100
        
        for e in all_exams:
            if e.id in all_scores_by_exam:
                totals = list(all_scores_by_exam[e.id].values())
                exam_trend_labels.append(e.name)
                exam_trend_avgs.append(round(sum(totals) / len(totals), 1))
                exam_trend_pass.append(round(sum(1 for t in totals if t >= full_total * 0.6) / len(totals) * 100, 1))
                exam_trend_excel.append(round(sum(1 for t in totals if t >= full_total * 0.85) / len(totals) * 100, 1))

    return jsonify({
        "code": 0,
        "data": {
            # 基础统计
            "total_students": total_students,
            "total_classes": total_classes,
            "total_teachers": total_teachers,

            # 成绩
            "score_stats": score_stats,
            "exam_trend": {
                "labels": exam_trend_labels,
                "avgs": exam_trend_avgs,
                "pass_rates": exam_trend_pass,
                "excel_rates": exam_trend_excel,
            },

            # 德育
            "discipline_count": discipline_count,
            "discipline_levels": disc_levels,
            "discipline_categories": discipline_by_category,

            # 考勤
            "attendance_rate": attendance_rate,
            "attendance_status": att_status,
            "leave_count": leave_count,
            "leave_approved": leave_approved,

            # 通知
            "notice_count": notice_count,
            "total_notices": total_notices,
            "notice_read_rate": notice_read_rate,

            # 家访
            "visit_count": visit_count,
            "visit_types": [{"name": vtype, "count": cnt}
                            for vtype, cnt in visit_by_type],

            # 五翼
            "wing_avg": wing_avg,
            "wing_dimensions": [{"name": dim, "score": round(float(s), 1)}
                                for dim, s in wing_dimensions],

            # 班级对比
            "class_scores": class_score_data,
            "class_discipline": class_disc_data,
            "class_attendance": class_att_data,

            # 趋势
            "trend": {
                "dates": trend_dates,
                "discipline": trend_discipline,
                "attendance": trend_attendance,
                "routine": trend_routine,
            },

            # 新增维度
            "mental_health": mh_risk,
            "risk_alerts": risk_stats,
            "risk_scan_date": risk_scan_date,
            "activity_stats": {
                "total_activities": total_activities,
                "reg_count": activity_reg_count,
            },
            "parent_meeting": {
                "total": pm_total,
                "signin_count": pm_signin_count,
                "rate": pm_rate,
            },
        }
    })


@cockpit_bp.route("/export/excel")
@login_required
@require_role("ms_admin", "grade_leader")
def export_excel():
    """导出驾驶舱数据为 Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"code": 1, "msg": "请先安装 openpyxl: pip install openpyxl"}), 500

    grade_id = request.args.get("grade_id", type=int) or session.get("grade_id")
    if not grade_id:
        g = Grade.query.order_by(Grade.sort_order).first()
        grade_id = g.id if g else 1

    days = request.args.get("days", 30, type=int)
    since = date.today() - timedelta(days=days)

    wb = openpyxl.Workbook()

    # 样式
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

    # ── Sheet 1: 概览 ──
    ws1 = wb.active
    ws1.title = "数据概览"
    ws1.append(["指标", "数值"])
    style_header(ws1, 1, 2)

    total_students = Student.query.filter_by(grade_id=grade_id, is_active=True).count()
    total_classes = Class.query.filter_by(grade_id=grade_id, is_active=True).count()
    disc_count = DisciplineRecord.query.filter(
        DisciplineRecord.grade_id == grade_id,
        DisciplineRecord.created_at >= since,
    ).count()

    overview_data = [
        ("在校学生", total_students),
        ("班级数", total_classes),
        ("违纪人次(近{}天)".format(days), disc_count),
        ("导出时间", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for label, val in overview_data:
        ws1.append([label, val])

    # ── Sheet 2: 成绩数据 ──
    ws2 = wb.create_sheet("成绩分析")
    # 批量预加载：一次查询所有考试的成绩 → 内存分组成绩统计
    all_scores = db.session.query(
        Score.exam_id, Score.student_id, func.sum(Score.score).label("total")
    ).join(Exam, Exam.id == Score.exam_id).filter(
        Exam.grade_id == grade_id,
        Score.score > 0
    ).group_by(Score.exam_id, Score.student_id).all()

    # 按考试分组聚合
    exam_scores_map = {}  # {exam_id: [total_score, ...]}
    for exam_id, stu_id, total in all_scores:
        exam_scores_map.setdefault(exam_id, []).append(float(total))

    exams = Exam.query.filter_by(grade_id=grade_id).order_by(Exam.exam_date.asc()).all()
    sc = Subject.query.count()  # 只查一次，不在循环内
    ws2.append(["考试名称", "日期", "均分", "及格率%", "优秀率%"])
    style_header(ws2, 1, 5)

    for e in exams:
        totals = exam_scores_map.get(e.id, [])
        if totals:
            ft = sc * 100
            avg = round(sum(totals) / len(totals), 1)
            pr = round(sum(1 for t in totals if t >= ft * 0.6) / len(totals) * 100, 1)
            er = round(sum(1 for t in totals if t >= ft * 0.85) / len(totals) * 100, 1)
            ws2.append([e.name, str(e.exam_date), avg, pr, er])

    # ── Sheet 3: 违纪明细 ──
    ws3 = wb.create_sheet("违纪记录")
    ws3.append(["姓名", "班级", "类型", "分类", "日期", "描述"])
    style_header(ws3, 1, 6)

    from sqlalchemy.orm import joinedload
    discs = DisciplineRecord.query.filter(
        DisciplineRecord.grade_id == grade_id,
        DisciplineRecord.created_at >= since,
    ).options(
        joinedload(DisciplineRecord.student).joinedload(Student.class_)
    ).order_by(DisciplineRecord.created_at.desc()).all()

    for d in discs:
        ws3.append([
            d.student.name if d.student else "",
            d.student.class_.name if d.student and d.student.class_ else "",
            d.type, d.category or "",
            d.created_at.strftime("%Y-%m-%d") if d.created_at else "",
            d.description or "",
        ])

    # ── Sheet 4: 考勤统计 ──
    ws4 = wb.create_sheet("考勤统计")
    classes = Class.query.filter_by(grade_id=grade_id, is_active=True).order_by(Class.name).all()
    ws4.append(["班级", "出勤", "迟到", "早退", "缺勤", "请假", "出勤率%"])
    style_header(ws4, 1, 7)

    for cls in classes:
        stats = db.session.query(
            Attendance.status,
            func.count(Attendance.id)
        ).filter(
            Attendance.class_id == cls.id,
            Attendance.record_date >= since,
        ).group_by(Attendance.status).all()
        stat_map = {s: c for s, c in stats}
        total = sum(stat_map.values())
        present = stat_map.get("present", 0)
        rate = round(present / total * 100, 1) if total > 0 else 0
        ws4.append([
            cls.name,
            stat_map.get("present", 0),
            stat_map.get("late", 0),
            stat_map.get("early", 0),
            stat_map.get("absent", 0),
            stat_map.get("leave", 0),
            rate,
        ])

    # 自动调整列宽
    for ws in [ws1, ws2, ws3, ws4]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"数据驾驶舱_{datetime.now().strftime('%Y%m%d')}.xlsx",
    )


@cockpit_bp.route("/export/pdf")
@login_required
@require_role("ms_admin", "grade_leader")
def export_pdf():
    """导出驾驶舱数据为 PDF（浏览器打印模式）"""
    grade_id = request.args.get("grade_id", type=int) or session.get("grade_id")
    days = request.args.get("days", 30, type=int)
    return render_template("cockpit/print.html",
                           grade_id=grade_id, days=days,
                           now=datetime.now())


# ══════════════════════════════════════════════════════════════
#  德育声呐战情大屏（Direction 4 — 流式可观测）
# ══════════════════════════════════════════════════════════════

def broadcast_sonar(data: dict):
    """向声呐大屏广播全局事件 — 供 push_event / 业务代码调用

    event_type 分类:
      discipline  — 违纪录入
      notify     — 家校通知推送
      risk       — AI 风险预警
      task       — 德育任务流转
      attendance — 考勤异常
      leave      — 请假审批
      general    — 通用消息
    """
    entry = {
        "event": data.get("type", "sonar"),
        "data": data,
        "ts": data.get("timestamp", time.time()),
    }

    # 1. 内存回放队列（始终写入，不依赖 Redis）
    _sonar_history.append(entry)
    if len(_sonar_history) > MAX_HISTORY:
        _sonar_history.pop(0)

    # 2. Redis Pub/Sub 广播（失败不影响内存队列）
    try:
        from blueprints.common import redis_client
        payload = json.dumps({
            "event": data.get("type", "sonar"),
            "data": data,
            "timestamp": time.time()
        }, ensure_ascii=False)
        redis_client.publish(SONAR_CHANNEL, payload)
    except Exception as e:
        from flask import current_app
        try:
            current_app.logger.warning(f"声呐 Redis 广播失败（已降级为内存模式）: {e}")
        except Exception:
            pass


@cockpit_bp.route("/sonar")
@login_required
def sonar():
    """德育声呐全屏战情大屏入口"""
    role = session.get("role", "")

    # 快照统计（页面加载时的即时数据）
    today = date.today()
    today_start = datetime.combine(today, datetime.min.time())

    stats = {
        "today_discipline": DisciplineRecord.query.filter(
            DisciplineRecord.created_at >= today_start).count(),
        "today_leave": LeaveRequest.query.filter(
            LeaveRequest.created_at >= today_start).count(),
        "risk_red": RiskRecord.query.filter(
            RiskRecord.risk_level == "red",
            RiskRecord.created_at >= today_start).count(),
        "total_students": Student.query.filter_by(is_active=True).count(),
    }

    return render_template("cockpit/sonar.html",
                           role=role, stats=stats,
                           static_css=url_for("static", filename="css"),
                           static_js=url_for("static", filename="js"))


@cockpit_bp.route("/api/sonar-stream")
@login_required
def sonar_stream():
    """SSE 流式推送 — Redis Pub/Sub 优先，自动降级为内存轮询"""

    def generate():
        from flask import current_app as app
        use_redis = True
        pubsub = None

        # 尝试 Redis Pub/Sub 连接
        try:
            from blueprints.common import redis_client
            pubsub = redis_client.pubsub()
            pubsub.subscribe(SONAR_CHANNEL)
            # 快速验证连接
            pubsub.get_message(timeout=1.0)
        except Exception as e:
            use_redis = False
            try:
                app.logger.warning(f"声呐 SSE Redis 不可用，降级为内存轮询: {e}")
            except Exception:
                pass

        t0 = time.time()
        last_ping = t0
        last_poll_idx = len(_sonar_history)  # 内存轮询起始位置

        try:
            while time.time() - t0 < MAX_SONAR_STREAM:
                # 心跳保活（15s 间隔）
                now = time.time()
                if now - last_ping > 15:
                    yield f": sonar-ping {time.time()}\n\n"
                    last_ping = now

                if use_redis and pubsub:
                    # Redis 模式：实时推送
                    try:
                        msg = pubsub.get_message(
                            ignore_subscribe_messages=True, timeout=0.5
                        )
                        if msg and msg.get("type") == "message":
                            yield f"data: {msg['data']}\n\n"
                    except Exception:
                        pass
                else:
                    # 内存降级模式：轮询内存队列
                    if len(_sonar_history) > last_poll_idx:
                        for entry in _sonar_history[last_poll_idx:]:
                            yield f"data: {json.dumps(entry, ensure_ascii=False)}\n\n"
                        last_poll_idx = len(_sonar_history)

                time.sleep(0.3)

            # 流结束前发送关闭信号
            yield "event: stream-end\ndata: {\"reason\":\"timeout\"}\n\n"
        except GeneratorExit:
            pass
        finally:
            if pubsub:
                try:
                    pubsub.unsubscribe(SONAR_CHANNEL)
                    pubsub.close()
                except Exception:
                    pass

    resp = Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
    )
    resp.headers["Cache-Control"] = "no-cache"
    resp.headers["X-Accel-Buffering"] = "no"
    return resp


@cockpit_bp.route("/api/sonar/history")
@login_required
def sonar_history():
    """返回内存中的最近事件（用于新连接快速补课）"""
    return jsonify({
        "code": 0,
        "data": list(_sonar_history[-MAX_HISTORY:])
    })


# ══════════════════════════════════════════════════════════════
#  AI 面谈战术简报 API（Phase 2 — 里子模块）
# ══════════════════════════════════════════════════════════════

@cockpit_bp.route('/api/sonar/generate-briefing/<int:student_id>', methods=['POST'])
@login_required
def sonar_generate_briefing(student_id):
    """
    声呐大屏专用：一键弹射 AI 面谈战术简报
    
    权限控制：
    - ms_admin（德育处）：可生成任意学生简报
    - grade_leader（年级组长）：可生成本年级学生简报
    - class_teacher（班主任）：仅可生成本班学生简报
    """
    from flask import current_app
    from models import Student
    
    # 1. 角色权限校验
    role = session.get("role")
    if role not in ["ms_admin", "grade_leader", "class_teacher"]:
        current_app.logger.warning(
            f"AI简报越权访问 attempt: user_id={session.get('user_id')}, role={role}"
        )
        return jsonify({"status": "error", "msg": "越权访问：无德育约谈权限"}), 403
    
    # 2. 学生档案存在性校验
    student = Student.query.get(student_id)
    if not student:
        return jsonify({"status": "error", "msg": "学生档案不存在"}), 404
    
    # 3. 班级/年级权限校验（班主任只能看本班学生）
    if role == "class_teacher":
        if student.class_id != session.get("class_id"):
            current_app.logger.warning(
                f"班主任越权访问学生档案: user_id={session.get('user_id')}, "
                f"student_id={student_id}, student_class={student.class_id}"
            )
            return jsonify({"status": "error", "msg": "无权访问该学生档案"}), 403
    
    # 4. 年级权限校验（年级组长只能看本年级学生）
    elif role == "grade_leader":
        if student.grade_id != session.get("grade_id"):
            current_app.logger.warning(
                f"年级组长越权访问学生档案: user_id={session.get('user_id')}, "
                f"student_id={student_id}, student_grade={student.grade_id}"
            )
            return jsonify({"status": "error", "msg": "无权访问该年级学生档案"}), 403
    
    # 5. 调用 AI 面谈简报生成器（懒加载，避免顶层导入断裂拖垮整个蓝图）
    try:
        from utils.interview_briefing import generate_student_briefing
    except ImportError as e:
        current_app.logger.error(f"AI 简报模块导入失败: {e}")
        return jsonify({"status": "error", "msg": "AI 简报模块暂不可用，请联系管理员"}), 503

    current_app.logger.info(f"开始生成 AI 面谈简报: student_id={student_id}, user_id={session.get('user_id')}")

    result = generate_student_briefing(student_id)
    
    if result["status"] == "error":
        current_app.logger.error(f"AI 面谈简报生成失败: student_id={student_id}, error={result.get('msg')}")
        return jsonify(result), 500
    
    current_app.logger.info(f"AI 面谈简报生成成功: student_id={student_id}, length={len(result.get('data', ''))}")
    
    # 6. 返回简报内容（Markdown 格式）
    return jsonify({
        "code": 0,
        "status": "success",
        "data": result["data"],
        "feature_vector": result.get("feature_vector", {}),
        "student_name": student.name,
        "class_name": student.class_.name if student.class_ else "未知"
    })
